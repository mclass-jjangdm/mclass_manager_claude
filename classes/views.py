from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db import models as django_models
import datetime
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter  # noqa: F401 (used in export view)

import calendar

from .models import Lesson, LessonSchedule, Enrollment, TuitionPayment, MonthlyEnrollment, DAY_CHOICES
from .forms import LessonForm, EnrollmentForm, TuitionPaymentForm, MonthlyEnrollmentEditForm, LessonTransferForm


def _assign_timetable_columns(items):
    """같은 요일의 수업 블록에 겹침 처리용 col/total_cols/left_pct/right_pct 할당"""
    if not items:
        return items

    items = sorted(items, key=lambda x: x['top'])
    lanes = []  # 각 lane의 현재 끝 위치(px)

    # 1단계: 그리디 lane 배정
    for item in items:
        start = item['top']
        end   = start + item['height']
        placed = False
        for i, lane_end in enumerate(lanes):
            if lane_end <= start:
                lanes[i] = end
                item['col'] = i
                placed = True
                break
        if not placed:
            item['col'] = len(lanes)
            lanes.append(end)

    # 2단계: 각 아이템이 겹치는 구간의 최대 col → total_cols 결정
    for item in items:
        s, e = item['top'], item['top'] + item['height']
        max_col = max(
            (o['col'] for o in items
             if not (o['top'] + o['height'] <= s or o['top'] >= e)),
            default=0,
        )
        item['total_cols'] = max_col + 1
        item['left_pct']   = item['col'] * 100 / item['total_cols']
        item['right_pct']  = (item['total_cols'] - item['col'] - 1) * 100 / item['total_cols']

    return items


def _get_lesson_form_context():
    """수업 폼에 필요한 교과/과목/교재 JSON 데이터 반환"""
    from subjects.models import Subject
    from bookstore.models import Book

    subjects = Subject.objects.filter(is_active=True).order_by('name')
    books = Book.objects.select_related('subject').order_by('title')

    subjects_data = [
        {'id': s.pk, 'name': s.name, 'category': s.category}
        for s in subjects
    ]
    books_data = [
        {'id': b.pk, 'title': b.title,
         'category': b.subject.category if b.subject else '기타'}
        for b in books
    ]

    categories = sorted(set(s['category'] for s in subjects_data))

    return {
        'categories': categories,
        'subjects_json': json.dumps(subjects_data, ensure_ascii=False),
        'books_json': json.dumps(books_data, ensure_ascii=False),
        'day_choices': DAY_CHOICES,
    }



def _save_schedules(post_data, lesson):
    """POST 데이터에서 요일별 시간 파싱 후 LessonSchedule 저장"""
    lesson.schedules.all().delete()
    for day, _ in DAY_CHOICES:
        if post_data.get(f'sched_{day}'):
            start = post_data.get(f'sched_{day}_start', '').strip()
            end = post_data.get(f'sched_{day}_end', '').strip()
            if start and end:
                LessonSchedule.objects.create(
                    lesson=lesson,
                    day=day,
                    start_time=start,
                    end_time=end,
                )


def _get_existing_schedules(lesson):
    """수업의 기존 일정을 {day: {start, end}} 형태로 반환"""
    result = {}
    for s in lesson.schedules.all():
        result[s.day] = {
            'start': s.start_time.strftime('%H:%M'),
            'end': s.end_time.strftime('%H:%M'),
        }
    return result


# ──────────────────────────────────────────────
# Lesson (수업)
# ──────────────────────────────────────────────

@login_required
def lesson_list(request):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    today = datetime.date.today()

    lessons = Lesson.objects.select_related('subject', 'teacher').prefetch_related('schedules')
    search = request.GET.get('search', '').strip()
    status = request.GET.get('status', 'active')

    # 이번 달 수업별 청구 금액 계산 (adjusted_tuition 합산)
    this_month_mes = list(
        MonthlyEnrollment.objects.filter(
            year=today.year,
            month=today.month,
        ).exclude(status='cancelled').select_related('lesson')
    )
    lesson_billing_map = {}
    for me in this_month_mes:
        lesson_billing_map[me.lesson_id] = lesson_billing_map.get(me.lesson_id, 0) + me.adjusted_tuition

    if search:
        lessons = lessons.filter(name__icontains=search)
    if status == 'active':
        lessons = lessons.filter(is_active=True)
    elif status == 'inactive':
        lessons = lessons.filter(is_active=False)

    # 청구 금액 속성 부여
    lessons = list(lessons)
    for lesson in lessons:
        lesson.current_billing = lesson_billing_map.get(lesson.pk, 0)

    # ── 시간표 데이터 (항상 활성 수업만, 검색 조건 무관) ──────────
    TIMETABLE_COLORS = [
        {'bg': '#e0e7ff', 'text': '#3730a3', 'border': '#c7d2fe'},  # indigo
        {'bg': '#ede9fe', 'text': '#6d28d9', 'border': '#ddd6fe'},  # violet
        {'bg': '#dbeafe', 'text': '#1e40af', 'border': '#bfdbfe'},  # blue
        {'bg': '#e0f2fe', 'text': '#0369a1', 'border': '#bae6fd'},  # sky
        {'bg': '#d1fae5', 'text': '#065f46', 'border': '#a7f3d0'},  # emerald
        {'bg': '#dcfce7', 'text': '#166534', 'border': '#bbf7d0'},  # green
        {'bg': '#fef9c3', 'text': '#854d0e', 'border': '#fde68a'},  # yellow
        {'bg': '#ffedd5', 'text': '#9a3412', 'border': '#fed7aa'},  # orange
        {'bg': '#fee2e2', 'text': '#991b1b', 'border': '#fecaca'},  # rose
        {'bg': '#fce7f3', 'text': '#9d174d', 'border': '#fbcfe8'},  # pink
        {'bg': '#f3e8ff', 'text': '#6b21a8', 'border': '#e9d5ff'},  # purple
        {'bg': '#ccfbf1', 'text': '#134e4a', 'border': '#99f6e4'},  # teal
    ]

    tt_lessons = list(
        Lesson.objects.filter(is_active=True)
        .select_related('subject', 'teacher')
        .prefetch_related('schedules', 'enrollments__student')
        .order_by('pk')
    )

    lesson_color_map = {
        lesson.pk: TIMETABLE_COLORS[i % len(TIMETABLE_COLORS)]
        for i, lesson in enumerate(tt_lessons)
    }
    tt_legend = [
        {'lesson': lesson, 'color': lesson_color_map[lesson.pk]}
        for lesson in tt_lessons
    ]

    # 시간 범위 계산
    all_starts, all_ends = [], []
    for lesson in tt_lessons:
        for s in lesson.schedules.all():
            all_starts.append(s.start_time.hour * 60 + s.start_time.minute)
            all_ends.append(s.end_time.hour * 60 + s.end_time.minute)

    if all_starts:
        base_min = (min(all_starts) // 60) * 60          # 내림하여 정시
        top_min  = ((max(all_ends) + 59) // 60) * 60     # 올림하여 정시
    else:
        base_min, top_min = 13 * 60, 22 * 60

    PX_PER_MIN = 1.2
    table_height = max(round((top_min - base_min) * PX_PER_MIN), 120)

    # 요일별 수업 블록 데이터
    day_map = {d: {'key': d, 'label': lbl, 'items': []} for d, lbl in DAY_CHOICES}
    for lesson in tt_lessons:
        color = lesson_color_map[lesson.pk]
        # 현재 활성 수강생 이름 목록
        students = sorted(
            e.student.name for e in lesson.enrollments.all() if e.is_active and e.student.is_active
        )
        for s in lesson.schedules.all():
            start_min = s.start_time.hour * 60 + s.start_time.minute
            end_min   = s.end_time.hour * 60 + s.end_time.minute
            day_map[s.day]['items'].append({
                'lesson':   lesson,
                'top':      round((start_min - base_min) * PX_PER_MIN),
                'height':   max(round((end_min - start_min) * PX_PER_MIN), 22),
                'time':     f"{s.start_time.strftime('%H:%M')}~{s.end_time.strftime('%H:%M')}",
                'color':    color,
                'students': students,
            })

    # 요일별 겹침 처리
    for day_data in day_map.values():
        day_data['items'] = _assign_timetable_columns(day_data['items'])

    timetable_days = [day_map[d] for d, _ in DAY_CHOICES]

    hour_labels = []
    h = base_min // 60
    while h * 60 <= top_min:
        hour_labels.append({
            'label': f'{h:02d}:00',
            'top':   round((h * 60 - base_min) * PX_PER_MIN),
        })
        h += 1

    return render(request, 'classes/lesson_list.html', {
        'lessons':        lessons,
        'search':         search,
        'status':         status,
        # timetable
        'timetable_days': timetable_days,
        'hour_labels':    hour_labels,
        'table_height':   table_height,
        'tt_legend':      tt_legend,
    })


@login_required
def lesson_detail(request, pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(
        Lesson.objects.prefetch_related('schedules', 'books'),
        pk=pk,
    )
    enrollments = lesson.enrollments.filter(student__is_active=True).select_related('student').prefetch_related('payments').order_by('-enrollment_date')

    return render(request, 'classes/lesson_detail.html', {
        'lesson': lesson,
        'enrollments': enrollments,
    })


@login_required
def lesson_create(request):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    if request.method == 'POST':
        form = LessonForm(request.POST)
        if form.is_valid():
            lesson = form.save()
            _save_schedules(request.POST, lesson)
            messages.success(request, f'수업 "{lesson.name}"이 생성되었습니다.')
            return redirect('classes:lesson_detail', pk=lesson.pk)
    else:
        form = LessonForm()  # teacher 기본값 null = 원장

    ctx = _get_lesson_form_context()
    ctx.update({
        'form': form,
        'action': 'create',
        'existing_schedules_json': '{}',
        'current_category': '',
    })
    return render(request, 'classes/lesson_form.html', ctx)


@login_required
def lesson_edit(request, pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(Lesson, pk=pk)

    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson)
        if form.is_valid():
            form.save()
            _save_schedules(request.POST, lesson)
            messages.success(request, f'수업 "{lesson.name}"이 수정되었습니다.')
            return redirect('classes:lesson_detail', pk=lesson.pk)
    else:
        form = LessonForm(instance=lesson)

    existing_schedules = _get_existing_schedules(lesson)
    current_category = lesson.subject.category if lesson.subject else ''

    ctx = _get_lesson_form_context()
    ctx.update({
        'form': form,
        'lesson': lesson,
        'action': 'edit',
        'existing_schedules_json': json.dumps(existing_schedules, ensure_ascii=False),
        'current_category': current_category,
    })
    return render(request, 'classes/lesson_form.html', ctx)


@login_required
def lesson_delete(request, pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(Lesson, pk=pk)

    if request.method == 'POST':
        name = lesson.name
        lesson.delete()
        messages.success(request, f'수업 "{name}"이 삭제되었습니다.')
        return redirect('classes:lesson_list')

    return render(request, 'classes/lesson_confirm_delete.html', {'lesson': lesson})


# ──────────────────────────────────────────────
# Enrollment (수강 신청)
# ──────────────────────────────────────────────

GRADE_LABELS = {
    'K5': '초5', 'K6': '초6',
    'K7': '중1', 'K8': '중2', 'K9': '중3',
    'K10': '고1', 'K11': '고2', 'K12': '고3',
}
GRADE_ORDER = ['K5', 'K6', 'K7', 'K8', 'K9', 'K10', 'K11', 'K12']


def _get_grade_groups(lesson):
    """학년별 학생 그룹 리스트와 이미 수강 중인 학생 ID 세트 반환.
    반환: (grade_groups, enrolled_ids)
    grade_groups = [{'key': 'K7', 'label': '중1', 'students': [...]}, ...]
    """
    from students.models import Student
    all_students = list(
        Student.objects.filter(quit_date__isnull=True).order_by('grade', 'name')
    )
    enrolled_ids = set(
        lesson.enrollments.filter(is_active=True).values_list('student_id', flat=True)
    )
    grade_groups = []
    for grade in GRADE_ORDER:
        group = [s for s in all_students if s.grade == grade]
        if group:
            grade_groups.append({
                'key': grade,
                'label': GRADE_LABELS.get(grade, grade),
                'students': group,
            })
    others = [s for s in all_students if s.grade not in GRADE_ORDER]
    if others:
        grade_groups.append({'key': 'other', 'label': '기타', 'students': others})
    return grade_groups, enrolled_ids


@login_required
def enrollment_create(request, pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(
        Lesson.objects.prefetch_related('schedules', 'enrollments'), pk=pk
    )

    if request.method == 'POST':
        student_pks = request.POST.getlist('student')
        enrollment_date = request.POST.get('enrollment_date', '').strip()
        end_date = request.POST.get('end_date', '').strip() or None
        tuition_adjustment = request.POST.get('tuition_adjustment', 0)
        memo = request.POST.get('memo', '')
        is_active = 'is_active' in request.POST

        errors = []
        if not student_pks:
            errors.append('학생을 한 명 이상 선택해 주세요.')
        if not enrollment_date:
            errors.append('수강 시작일을 입력해 주세요.')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            from students.models import Student
            success_count = 0
            for spk in student_pks:
                try:
                    student = Student.objects.get(pk=spk)
                    Enrollment.objects.create(
                        student=student,
                        lesson=lesson,
                        enrollment_date=enrollment_date,
                        end_date=end_date,
                        tuition_adjustment=tuition_adjustment,
                        memo=memo,
                        is_active=is_active,
                    )
                    success_count += 1
                except Exception:
                    messages.warning(request, f'{student.name} 학생은 이미 수강 중입니다.')
            if success_count:
                messages.success(request, f'{success_count}명의 수강 신청이 완료되었습니다.')
            return redirect('classes:lesson_detail', pk=lesson.pk)

    grade_groups, enrolled_ids = _get_grade_groups(lesson)
    today = datetime.date.today()

    return render(request, 'classes/enrollment_form.html', {
        'lesson': lesson,
        'action': 'create',
        'grade_groups': grade_groups,
        'enrolled_ids': enrolled_ids,
        'today': today,
        'prefill_end_date': lesson.end_date if lesson.is_special else None,
        'prefill_start_date': lesson.start_date if lesson.is_special else None,
    })


@login_required
def enrollment_edit(request, pk, enroll_pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(Lesson.objects.prefetch_related('schedules'), pk=pk)
    enrollment = get_object_or_404(Enrollment, pk=enroll_pk, lesson=lesson)

    if request.method == 'POST':
        form = EnrollmentForm(request.POST, instance=enrollment)
        if form.is_valid():
            form.save()
            messages.success(request, '수강 신청 정보가 수정되었습니다.')
            return redirect('classes:lesson_detail', pk=lesson.pk)
    else:
        form = EnrollmentForm(instance=enrollment)

    return render(request, 'classes/enrollment_form.html', {
        'form': form,
        'lesson': lesson,
        'enrollment': enrollment,
        'action': 'edit',
    })


@login_required
def enrollment_delete(request, pk, enroll_pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    lesson = get_object_or_404(Lesson, pk=pk)
    enrollment = get_object_or_404(Enrollment, pk=enroll_pk, lesson=lesson)

    if request.method == 'POST':
        student_name = enrollment.student.name

        # 납부 기록이 있으면 삭제 불가 (TuitionPayment.on_delete=CASCADE로 납부 내역까지 사라짐)
        if enrollment.payments.exists():
            messages.error(
                request,
                f'{student_name} 학생은 납부 기록이 있어 수강을 삭제할 수 없습니다. '
                '비활성화(종료 처리)를 이용해 주세요.'
            )
            return redirect('classes:lesson_detail', pk=lesson.pk)

        # 연결된 MonthlyEnrollment를 cancelled 처리 (고아 레코드 방지)
        MonthlyEnrollment.objects.filter(
            student=enrollment.student,
            lesson=enrollment.lesson,
        ).update(status='cancelled')

        enrollment.delete()
        messages.success(request, f'{student_name} 학생의 수강이 취소되었습니다.')
        return redirect('classes:lesson_detail', pk=lesson.pk)

    return render(request, 'classes/enrollment_confirm_delete.html', {
        'enrollment': enrollment,
        'lesson': lesson,
    })


# ──────────────────────────────────────────────
# TuitionPayment (수강료 납부)
# ──────────────────────────────────────────────

@login_required
def tuition_payment_create(request, enroll_pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    enrollment = get_object_or_404(Enrollment, pk=enroll_pk)
    today = datetime.date.today()

    if request.method == 'POST':
        form = TuitionPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.enrollment = enrollment
            payment.save()
            messages.success(request, f'{payment.year}년 {payment.month}월 수강료가 납부 처리되었습니다.')
            return redirect('classes:lesson_detail', pk=enrollment.lesson.pk)
    else:
        form = TuitionPaymentForm(initial={
            'year': today.year,
            'month': today.month,
            'amount': enrollment.adjusted_tuition,
            'payment_date': today,
        })

    return render(request, 'classes/tuition_payment_form.html', {
        'form': form,
        'enrollment': enrollment,
    })


@login_required
def tuition_payment_delete(request, enroll_pk, pay_pk):
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    enrollment = get_object_or_404(Enrollment, pk=enroll_pk)
    payment = get_object_or_404(TuitionPayment, pk=pay_pk, enrollment=enrollment)

    if request.method == 'POST':
        lesson_pk = enrollment.lesson.pk
        payment.delete()
        messages.success(request, '납부 기록이 삭제되었습니다.')
        return redirect('classes:lesson_detail', pk=lesson_pk)

    return render(request, 'classes/tuition_payment_confirm_delete.html', {
        'payment': payment,
        'enrollment': enrollment,
    })


# ──────────────────────────────────────────────
# MonthlyEnrollment (월별 수강 신청)
# ──────────────────────────────────────────────

def _get_target_month(request, param_year='year', param_month='month'):
    """쿼리 파라미터에서 연/월을 읽고, 없으면 다음 달 반환."""
    today = datetime.date.today()
    if today.month == 12:
        default_year, default_month = today.year + 1, 1
    else:
        default_year, default_month = today.year, today.month + 1
    try:
        target_year = int(request.GET.get(param_year, default_year))
        target_month = int(request.GET.get(param_month, default_month))
        if not (1 <= target_month <= 12):
            target_year, target_month = default_year, default_month
    except (ValueError, TypeError):
        target_year, target_month = default_year, default_month
    return target_year, target_month, default_year, default_month


def _prev_next_month(year, month):
    """이전 달 / 다음 달 (year, month) 튜플 반환."""
    if month == 1:
        prev = (year - 1, 12)
    else:
        prev = (year, month - 1)
    if month == 12:
        nxt = (year + 1, 1)
    else:
        nxt = (year, month + 1)
    return prev, nxt


@login_required
def monthly_enrollment_list(request):
    """월별 수강 신청 현황 목록"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    target_year, target_month, default_year, default_month = _get_target_month(request)
    status_filter = request.GET.get('status', '')

    qs = MonthlyEnrollment.objects.filter(
        year=target_year, month=target_month,
    ).select_related('student', 'lesson', 'lesson__teacher', 'lesson__subject')

    if status_filter in ('pending', 'confirmed', 'cancelled'):
        qs = qs.filter(status=status_filter)

    # 수업별 그룹핑
    lessons_dict = {}
    for me in qs:
        lid = me.lesson_id
        if lid not in lessons_dict:
            lessons_dict[lid] = {'lesson': me.lesson, 'items': []}
        lessons_dict[lid]['items'].append(me)
    lesson_groups = list(lessons_dict.values())

    total = qs.count()
    pending_count   = qs.filter(status='pending').count()
    confirmed_count = qs.filter(status='confirmed').count()
    cancelled_count = qs.filter(status='cancelled').count()

    (prev_year, prev_month), (next_year_nav, next_month_nav) = _prev_next_month(target_year, target_month)

    return render(request, 'classes/monthly_enrollment_list.html', {
        'target_year': target_year,
        'target_month': target_month,
        'default_year': default_year,
        'default_month': default_month,
        'lesson_groups': lesson_groups,
        'total': total,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'cancelled_count': cancelled_count,
        'status_filter': status_filter,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year_nav': next_year_nav,
        'next_month_nav': next_month_nav,
    })


@login_required
def monthly_enrollment_create(request):
    """현재 활성 Enrollment 기반으로 MonthlyEnrollment 일괄 생성 (pending)"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    today = datetime.date.today()
    if today.month == 12:
        default_year, default_month = today.year + 1, 1
    else:
        default_year, default_month = today.year, today.month + 1

    if request.method == 'POST':
        try:
            target_year  = int(request.POST.get('target_year',  default_year))
            target_month = int(request.POST.get('target_month', default_month))
        except (ValueError, TypeError):
            target_year, target_month = default_year, default_month

        selected_ids = request.POST.getlist('enrollment_ids')
        created_count = 0
        skip_count    = 0

        for enroll_id in selected_ids:
            try:
                enroll = Enrollment.objects.select_related('student', 'lesson').get(pk=enroll_id)
                is_enrollment_month = (
                    enroll.enrollment_date.year == target_year
                    and enroll.enrollment_date.month == target_month
                )
                obj, created = MonthlyEnrollment.objects.get_or_create(
                    student=enroll.student,
                    lesson=enroll.lesson,
                    year=target_year,
                    month=target_month,
                    defaults={
                        'status': 'pending',
                        'tuition_fee': enroll.lesson.base_tuition,
                        'tuition_adjustment': enroll.tuition_adjustment if is_enrollment_month else 0,
                        'memo': '',
                    },
                )
                if created:
                    created_count += 1
                else:
                    skip_count += 1
            except Enrollment.DoesNotExist:
                continue

        msg = f'{target_year}년 {target_month}월 수강 신청 {created_count}건 생성 완료'
        if skip_count:
            msg += f' (이미 존재 {skip_count}건 제외)'
        messages.success(request, msg)
        return redirect(f"{request.path}?year={target_year}&month={target_month}")

    # GET
    target_year, target_month, default_year, default_month = _get_target_month(request)

    first_of_target = datetime.date(target_year, target_month, 1)

    active_enrollments = Enrollment.objects.filter(
        is_active=True,
        student__is_active=True,
        lesson__is_special=False,
    ).filter(
        django_models.Q(end_date__isnull=True) | django_models.Q(end_date__gte=first_of_target)
    ).select_related(
        'student', 'lesson', 'lesson__teacher', 'lesson__subject'
    ).order_by('lesson__name', 'student__name')

    # 이미 해당 월에 MonthlyEnrollment 존재하는 (student_id, lesson_id) 쌍
    already_set = set(
        MonthlyEnrollment.objects.filter(
            year=target_year, month=target_month,
        ).values_list('student_id', 'lesson_id')
    )

    lessons_dict = {}
    total_new = 0
    total_exists = 0
    for enroll in active_enrollments:
        lid = enroll.lesson_id
        if lid not in lessons_dict:
            lessons_dict[lid] = {
                'lesson': enroll.lesson,
                'items': [],
                'new_count': 0,
                'exists_count': 0,
            }
        already = (enroll.student_id, enroll.lesson_id) in already_set
        lessons_dict[lid]['items'].append({'enrollment': enroll, 'already': already})
        if already:
            lessons_dict[lid]['exists_count'] += 1
            total_exists += 1
        else:
            lessons_dict[lid]['new_count'] += 1
            total_new += 1

    lesson_groups = list(lessons_dict.values())
    (prev_year, prev_month), (next_year_nav, next_month_nav) = _prev_next_month(target_year, target_month)

    return render(request, 'classes/monthly_enrollment_create.html', {
        'target_year': target_year,
        'target_month': target_month,
        'default_year': default_year,
        'default_month': default_month,
        'lesson_groups': lesson_groups,
        'total_new': total_new,
        'total_exists': total_exists,
        'total': total_new + total_exists,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year_nav': next_year_nav,
        'next_month_nav': next_month_nav,
    })


@login_required
def monthly_enrollment_edit(request, pk):
    """월별 수강 신청 개별 수정 (수업 변경 포함)"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    me = get_object_or_404(
        MonthlyEnrollment.objects.select_related('student', 'lesson'),
        pk=pk,
    )

    if request.method == 'POST':
        form = MonthlyEnrollmentEditForm(request.POST, instance=me)
        if form.is_valid():
            new_lesson = form.cleaned_data['lesson']
            # 수업이 변경되는 경우 unique_together 충돌 체크
            if new_lesson != me.lesson:
                if MonthlyEnrollment.objects.filter(
                    student=me.student,
                    lesson=new_lesson,
                    year=me.year,
                    month=me.month,
                ).exclude(pk=me.pk).exists():
                    messages.error(
                        request,
                        f'{me.student.name} 학생은 이미 [{new_lesson.name}] 수업에 {me.year}년 {me.month}월 수강 신청되어 있습니다.'
                    )
                    return redirect('classes:monthly_enrollment_edit', pk=pk)
            form.save()
            messages.success(request, f'{me.student.name} 학생의 수강 신청이 수정되었습니다.')
            return redirect(f"/classes/monthly/?year={me.year}&month={me.month}")
    else:
        form = MonthlyEnrollmentEditForm(instance=me)

    return render(request, 'classes/monthly_enrollment_edit.html', {
        'form': form,
        'me': me,
    })


@login_required
def monthly_enrollment_delete(request, pk):
    """월별 수강 신청 개별 삭제 (Enrollment 이력 보존)"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    me = get_object_or_404(
        MonthlyEnrollment.objects.select_related('student', 'lesson'),
        pk=pk,
    )

    if request.method == 'POST':
        year, month = me.year, me.month
        student_name = me.student.name
        me.delete()
        messages.success(request, f'{student_name} 학생의 {year}년 {month}월 수강 신청이 삭제되었습니다.')
        return redirect(f"/classes/monthly/?year={year}&month={month}")

    return render(request, 'classes/monthly_enrollment_confirm_delete.html', {'me': me})


@login_required
def monthly_enrollment_bulk_confirm(request):
    """선택된 MonthlyEnrollment를 일괄 확정 (status → confirmed)"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    if request.method == 'POST':
        pks = request.POST.getlist('monthly_ids')
        year  = request.POST.get('target_year')
        month = request.POST.get('target_month')
        count = MonthlyEnrollment.objects.filter(pk__in=pks).update(status='confirmed')
        messages.success(request, f'{count}건의 수강 신청이 확정되었습니다.')
        redirect_url = '/classes/monthly/'
        if year and month:
            redirect_url += f'?year={year}&month={month}'
        return redirect(redirect_url)
    return redirect('classes:monthly_enrollment_list')


@login_required
def monthly_enrollment_bulk_delete(request):
    """선택된 MonthlyEnrollment를 일괄 삭제"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    if request.method == 'POST':
        pks = request.POST.getlist('monthly_ids')
        year  = request.POST.get('target_year')
        month = request.POST.get('target_month')
        count, _ = MonthlyEnrollment.objects.filter(pk__in=pks).delete()
        messages.success(request, f'{count}건의 수강 신청이 삭제되었습니다.')
        redirect_url = '/classes/monthly/'
        if year and month:
            redirect_url += f'?year={year}&month={month}'
        return redirect(redirect_url)
    return redirect('classes:monthly_enrollment_list')


@login_required
def monthly_enrollment_export_xlsx(request):
    """월별 수강 신청 현황 엑셀 다운로드"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    target_year, target_month, _, _ = _get_target_month(request)
    status_filter = request.GET.get('status', '')

    qs = MonthlyEnrollment.objects.filter(
        year=target_year, month=target_month,
    ).select_related('student', 'lesson')

    if status_filter in ('pending', 'confirmed', 'cancelled'):
        qs = qs.filter(status=status_filter)

    qs = qs.order_by('lesson__name', 'student__name')

    # ── 워크북 생성 ───────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{target_year}년 {target_month}월'

    # 스타일 정의
    header_font = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
    cell_font   = Font(name='맑은 고딕', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')
    right_align  = Alignment(horizontal='right',  vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 헤더 행 ───────────────────────────────────────────
    headers = ['이름', '부모 전화번호', '청구금액', '내용', '메모']
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 20

    # ── 컬럼 너비 설정 ────────────────────────────────────
    col_widths = [12, 16, 12, 24, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── 데이터 행 ─────────────────────────────────────────
    for row_num, me in enumerate(qs, start=2):
        row_data = [
            me.student.name,
            me.student.get_formatted_parent_phone() or '',
            me.adjusted_tuition,
            me.lesson.name,
            me.memo,
        ]
        ws.append(row_data)
        for col_idx, _ in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border
            if col_idx == 3:
                cell.alignment = right_align
                cell.number_format = '#,##0'
            else:
                cell.alignment = left_align
        ws.row_dimensions[row_num].height = 17

    # ── 응답 반환 ─────────────────────────────────────────
    STATUS_LABEL = {'pending': '신청 중', 'confirmed': '확정', 'cancelled': '취소'}
    status_suffix = f'_{STATUS_LABEL[status_filter]}' if status_filter else ''
    filename = f'{target_year}년{target_month}월_수강신청{status_suffix}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
# 반 이동 (Lesson Transfer)
# ──────────────────────────────────────────────

@login_required
def lesson_transfer(request):
    """학생의 반 이동 처리: 기존 수강 종료 + 신규 수강 시작 + 수강료 일할 계산"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        return redirect('index')

    preview = None

    if request.method == 'POST':
        form = LessonTransferForm(request.POST)
        if form.is_valid():
            student     = form.cleaned_data['student']
            from_lesson = form.cleaned_data['from_lesson']
            to_lesson   = form.cleaned_data['to_lesson']
            transfer_date = form.cleaned_data['transfer_date']

            year  = transfer_date.year
            month = transfer_date.month
            days_in_month = calendar.monthrange(year, month)[1]

            # 전반일 이전 일수 (구 반 수강일): 1일 ~ 전반일-1일
            old_days = transfer_date.day - 1
            # 전반일부터 월말 일수 (신 반 수강일): 전반일 ~ 말일
            new_days = days_in_month - transfer_date.day + 1

            old_tuition = round(from_lesson.base_tuition * old_days / days_in_month)
            new_tuition = round(to_lesson.base_tuition   * new_days / days_in_month)

            # 미리보기 요청 (confirm=0)
            if request.POST.get('confirm') != '1':
                preview = {
                    'student': student,
                    'from_lesson': from_lesson,
                    'to_lesson': to_lesson,
                    'transfer_date': transfer_date,
                    'year': year,
                    'month': month,
                    'days_in_month': days_in_month,
                    'old_days': old_days,
                    'new_days': new_days,
                    'old_tuition': old_tuition,
                    'new_tuition': new_tuition,
                }
                return render(request, 'classes/lesson_transfer.html', {
                    'form': form,
                    'preview': preview,
                })

            # ── 확정 처리 ──────────────────────────────────────

            # 1. 기존 Enrollment 비활성화
            old_enrollment = Enrollment.objects.get(student=student, lesson=from_lesson, is_active=True)
            old_enrollment.is_active = False
            old_enrollment.end_date  = transfer_date - datetime.timedelta(days=1)
            old_enrollment.save()

            # 2. 기존 반 이번 달 MonthlyEnrollment 수강료 조정
            try:
                old_me = MonthlyEnrollment.objects.get(
                    student=student, lesson=from_lesson, year=year, month=month
                )
                old_me.tuition_fee = old_tuition
                old_me.tuition_adjustment = 0
                old_me.memo = (old_me.memo + f'\n전반({transfer_date} 이전 {old_days}일 분)').strip()
                old_me.save()
            except MonthlyEnrollment.DoesNotExist:
                # 해당 월 수강 신청이 없으면 생성하지 않음 (이미 없는 달이므로 무시)
                pass

            # 3. 신규 Enrollment 생성 또는 재활성화
            new_enrollment, created = Enrollment.objects.get_or_create(
                student=student,
                lesson=to_lesson,
                defaults={
                    'enrollment_date': transfer_date,
                    'is_active': True,
                },
            )
            if not created:
                new_enrollment.is_active = True
                new_enrollment.enrollment_date = transfer_date
                new_enrollment.end_date = None
                new_enrollment.save()

            # 4. 신규 반 이번 달 MonthlyEnrollment 생성 또는 갱신
            new_me, me_created = MonthlyEnrollment.objects.get_or_create(
                student=student,
                lesson=to_lesson,
                year=year,
                month=month,
                defaults={
                    'status': 'pending',
                    'tuition_fee': new_tuition,
                    'tuition_adjustment': 0,
                    'memo': f'전반({transfer_date}부터 {new_days}일 분)',
                },
            )
            if not me_created:
                new_me.tuition_fee = new_tuition
                new_me.tuition_adjustment = 0
                new_me.memo = (new_me.memo + f'\n전반({transfer_date}부터 {new_days}일 분)').strip()
                new_me.save()

            messages.success(
                request,
                f'{student.name} 학생이 [{from_lesson.name}] → [{to_lesson.name}]으로 전반되었습니다. '
                f'({year}년 {month}월 수강료: {old_tuition:,}원 / {new_tuition:,}원)'
            )
            return redirect(f'/classes/monthly/?year={year}&month={month}')
    else:
        form = LessonTransferForm()

    return render(request, 'classes/lesson_transfer.html', {
        'form': form,
        'preview': preview,
    })
