import logging
import datetime
import csv
from django.db import models
import io
from openpyxl import load_workbook, Workbook
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic.edit import CreateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import ListView
from django.http import HttpResponse
from .models import Student, School, StudentFile
from .forms import StudentForm, StudentImportForm, BulkSMSForm, StudentFileUploadForm
from common.utils import send_sms, validate_uploaded_file

logger = logging.getLogger(__name__)


class StudentListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'students/student_list.html'
    success_url = reverse_lazy('students:student_list')
    context_object_name = 'students'

    def get_queryset(self):
        from django.db.models import Sum, F, Case, When, IntegerField
        from bookstore.models import BookSale

        queryset = Student.objects.select_related('school')
        search_query = self.request.GET.get('search', '')
        show_inactive = self.request.GET.get('show_inactive') == 'on'

        if search_query:
            queryset = queryset.filter(name__icontains=search_query)
        if not show_inactive:
            queryset = queryset.filter(is_active=True)

        # 미결제 도서 금액 총액을 annotate로 추가
        queryset = queryset.annotate(
            unpaid_book_total=Sum(
                Case(
                    When(book_sales__is_paid=False,
                         then=F('book_sales__price') * F('book_sales__quantity')),
                    default=0,
                    output_field=IntegerField()
                )
            )
        )

        return queryset.order_by('-is_active', 'grade', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['show_inactive'] = self.request.GET.get('show_inactive') == 'on'
        context['search_query'] = self.request.GET.get('search', '')

        # 정렬 방식 가져오기 (기본값: grade)
        group_by = self.request.GET.get('group_by', 'grade')
        context['group_by'] = group_by

        # 수강료 미납 금액 계산 (학생별) - MonthlyEnrollment 기준
        from classes.models import MonthlyEnrollment, TuitionPayment
        from django.utils import timezone as tz
        from django.db.models import Q as dQ
        today_for_tuition = tz.now().date()

        # 모든 납부 기록 (student_id, lesson_id, year, month)
        all_paid_quads = set(
            TuitionPayment.objects.values_list(
                'enrollment__student_id', 'enrollment__lesson_id', 'year', 'month'
            )
        )
        all_mes = MonthlyEnrollment.objects.exclude(
            status='cancelled'
        ).select_related('lesson')

        unpaid_tuition_dict = {}
        for me in all_mes:
            if (me.student_id, me.lesson_id, me.year, me.month) not in all_paid_quads:
                unpaid_tuition_dict[me.student_id] = unpaid_tuition_dict.get(me.student_id, 0) + me.adjusted_tuition

        # 학생 그룹화
        students = self.get_queryset()

        # 학생 객체에 unpaid_tuition 속성 부여
        for student in students:
            student.unpaid_tuition = unpaid_tuition_dict.get(student.pk, 0)
        from collections import defaultdict

        if group_by == 'school':
            # 학교별 그룹화
            grouped = defaultdict(list)
            for student in students:
                school_name = student.school.name if student.school else '미지정'
                grouped[school_name].append(student)
            context['grouped_students'] = dict(sorted(grouped.items()))
        else:
            # 학년별 그룹화 (기본)
            grade_order = ['K5', 'K6', 'K7', 'K8', 'K9', 'K10', 'K11', 'K12', '졸업']
            grouped = defaultdict(list)
            for student in students:
                grade = student.grade if student.grade else '미지정'
                grouped[grade].append(student)

            # 학년 순서대로 정렬
            sorted_grouped = {}
            for grade in grade_order:
                if grade in grouped:
                    sorted_grouped[grade] = grouped[grade]
            # 미지정이나 기타 학년 추가
            for grade, student_list in grouped.items():
                if grade not in grade_order:
                    sorted_grouped[grade] = student_list

            context['grouped_students'] = sorted_grouped

        return context


class StudentCreateView(LoginRequiredMixin, CreateView):
    form_class = StudentForm
    template_name = 'students/student_form.html'
    success_url = reverse_lazy('students:student_list')

    def get_initial(self):
        initial = super().get_initial()
        consultation_pk = self.request.GET.get('consultation_pk')
        if consultation_pk:
            try:
                from homepage.models import ConsultationRequest
                c = ConsultationRequest.objects.get(pk=consultation_pk)
                initial['name'] = c.name
                initial['gender'] = c.gender
                initial['grade'] = c.grade
                initial['phone_number'] = c.phone_number
                initial['email'] = c.email
                initial['parent_phone'] = c.parent_phone
                initial['interview_info'] = c.get_interview_info_text()
            except Exception:
                pass
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['consultation_pk'] = self.request.GET.get('consultation_pk')
        return ctx

    def form_valid(self, form):
        student = form.save(commit=False)
        student.student_id = self.generate_student_id()
        student.save()

        # 상담 신청이 있으면 처리 상태를 '등록 완료'로 업데이트
        consultation_pk = self.request.POST.get('consultation_pk') or self.request.GET.get('consultation_pk')
        if consultation_pk:
            try:
                from homepage.models import ConsultationRequest
                c = ConsultationRequest.objects.get(pk=consultation_pk)
                c.status = 'registered'
                c.save()
            except Exception:
                pass

        messages.success(self.request, f"학생 '{student.name}' 등록되었습니다.")
        return redirect('students:student_list')

    def generate_student_id(self):
        import random
        return ''.join([str(random.randint(0, 9)) for _ in range(8)])


@login_required
def student_detail(request, pk):
    from bookstore.models import BookSale
    from grades.models import Grade
    from django.utils import timezone

    student = get_object_or_404(Student, pk=pk)

    # 교재 구매 내역 조회
    unpaid_sales = BookSale.objects.filter(student=student, is_paid=False).select_related('book').order_by('-sale_date', '-pk')
    paid_sales = BookSale.objects.filter(student=student, is_paid=True).select_related('book').order_by('-sale_date', '-pk')

    # 교재 대금 계산
    total_unpaid_books = sum(sale.get_total_price() for sale in unpaid_sales)
    total_paid_books = sum(sale.get_total_price() for sale in paid_sales)

    # 수강료 계산 - MonthlyEnrollment 기반
    from classes.models import MonthlyEnrollment, Enrollment, TuitionPayment
    from django.db.models import Sum, Q as dQ
    today = timezone.now().date()

    # 활성 Enrollment (하단 수강 테이블용) — end_date 지난 것 제외
    active_enrollments = Enrollment.objects.filter(
        student=student, is_active=True,
    ).filter(
        models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
    ).select_related('lesson')
    # 납부 버튼용 - 비활성 수업도 포함 (MonthlyEnrollment가 있으면 납부 가능해야 함)
    enrollment_by_lesson = {e.lesson_id: e for e in Enrollment.objects.filter(student=student).select_related('lesson')}

    # 당월 MonthlyEnrollment
    current_mes = MonthlyEnrollment.objects.filter(
        student=student, year=today.year, month=today.month,
    ).exclude(status='cancelled').select_related('lesson')
    current_month_tuition = sum(me.adjusted_tuition for me in current_mes)

    # 당월 납부 완료 lesson_id 집합
    paid_this_month_lesson_ids = set(
        TuitionPayment.objects.filter(
            enrollment__student=student, year=today.year, month=today.month,
        ).values_list('enrollment__lesson_id', flat=True)
    )

    # 당월 수업별 항목 (ME + 납부상태 + enrollment)
    current_me_items = [
        {
            'me': me,
            'enrollment': enrollment_by_lesson.get(me.lesson_id),
            'is_paid': me.lesson_id in paid_this_month_lesson_ids,
        }
        for me in current_mes
    ]

    # 과거 미납 (이번 달 이전 MonthlyEnrollment)
    paid_quads = set(
        TuitionPayment.objects.filter(
            enrollment__student=student
        ).values_list('enrollment__lesson_id', 'year', 'month')
    )
    past_mes = MonthlyEnrollment.objects.filter(
        student=student,
    ).filter(
        dQ(year__lt=today.year) | dQ(year=today.year, month__lt=today.month)
    ).exclude(status='cancelled').select_related('lesson')

    past_unpaid_mes = [
        me for me in past_mes
        if (me.lesson_id, me.year, me.month) not in paid_quads
    ]
    past_unpaid_tuition = sum(me.adjusted_tuition for me in past_unpaid_mes)

    unpaid_lesson_ids = {me.lesson_id for me in past_unpaid_mes}
    enrollments_for_unpaid = {
        e.lesson_id: e
        for e in Enrollment.objects.filter(student=student, lesson_id__in=unpaid_lesson_ids)
    } if unpaid_lesson_ids else {}

    past_unpaid_me_items = sorted(
        [
            {
                'me': me,
                'enrollment': enrollments_for_unpaid.get(me.lesson_id),
            }
            for me in past_unpaid_mes
        ],
        key=lambda x: (x['me'].year, x['me'].month),
        reverse=True,
    )

    # 특별 수업 수강 현황
    special_enrollments = Enrollment.objects.filter(
        student=student,
        is_active=True,
        lesson__is_special=True,
    ).select_related('lesson')

    paid_special_enrollment_ids = set(
        TuitionPayment.objects.filter(
            enrollment__student=student,
            enrollment__lesson__is_special=True,
        ).values_list('enrollment_id', flat=True)
    )

    special_enrollment_items = [
        {
            'enrollment': e,
            'is_paid': e.pk in paid_special_enrollment_ids,
        }
        for e in special_enrollments
    ]

    # 납부 이력
    tuition_payments = TuitionPayment.objects.filter(
        enrollment__student=student
    ).select_related('enrollment__lesson').order_by('-year', '-month', '-id')
    total_tuition_paid = tuition_payments.aggregate(total=Sum('amount'))['total'] or 0

    # 환불 이력
    from classes.models import WithdrawalRefund
    withdrawal_refunds = WithdrawalRefund.objects.filter(
        student=student
    ).select_related('enrollment__lesson').order_by('-year', '-month', '-id')
    total_tuition_refund = withdrawal_refunds.aggregate(total=Sum('refund_amount'))['total'] or 0
    total_tuition_net = total_tuition_paid - total_tuition_refund

    # 납부+환불 통합 정렬 (최신순)
    combined_tuition_history = sorted(
        [{'type': 'payment', 'year': tp.year, 'month': tp.month, 'obj': tp} for tp in tuition_payments] +
        [{'type': 'refund',  'year': wr.year, 'month': wr.month, 'obj': wr} for wr in withdrawal_refunds],
        key=lambda x: (x['year'], x['month'], 1 if x['type'] == 'refund' else 0, x['obj'].id),
        reverse=True,
    )

    # 성적 데이터 조회
    internal_grades = Grade.objects.filter(
        student=student,
        grade_type='internal'
    ).select_related('subject').order_by('-year', '-semester', 'subject__subject_code')

    mock_grades = Grade.objects.filter(
        student=student,
        grade_type='mock'
    ).select_related('subject').order_by('-exam_year', '-exam_month', 'subject__subject_code')

    # 학기별 평균 내신 등급 계산 (진로선택 과목 제외)
    from collections import defaultdict
    from decimal import Decimal
    import json

    semester_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})
    year_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})

    # 학기별/교과별 성적 데이터 (차트용) - 진로선택 제외, 교과별 집계
    semester_category_grades = defaultdict(lambda: defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0}))

    # 교과별 전체 통계 (교과 조합 분석용)
    category_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})

    for grade in internal_grades:
        # 등급/단위수가 없는 경우(중학교 성적 등) 통계 계산에서 제외
        if grade.grade_rank is None or grade.credits is None:
            continue

        # 차트용 데이터 수집 (진로선택 과목 제외)
        if not grade.is_elective:
            key = f"{grade.year}-{grade.semester}"
            category = grade.curriculum or '기타'
            semester_category_grades[key][category]['total_weighted'] += grade.grade_rank * grade.credits
            semester_category_grades[key][category]['total_credits'] += grade.credits

            # 교과별 전체 통계
            category_stats[category]['total_weighted'] += grade.grade_rank * grade.credits
            category_stats[category]['total_credits'] += grade.credits

        if grade.is_elective:  # 진로선택 과목은 평균 계산에서 제외
            continue

        semester_key = (grade.year, grade.semester)
        semester_stats[semester_key]['total_weighted'] += grade.grade_rank * grade.credits
        semester_stats[semester_key]['total_credits'] += grade.credits

        # 학년별 통계
        year_stats[grade.year]['total_weighted'] += grade.grade_rank * grade.credits
        year_stats[grade.year]['total_credits'] += grade.credits

    # 학기별 평균 계산 및 정렬
    semester_averages = []
    for (year, semester), stats in sorted(semester_stats.items()):
        if stats['total_credits'] > 0:
            avg = Decimal(stats['total_weighted']) / Decimal(stats['total_credits'])
            semester_averages.append({
                'year': year,
                'semester': semester,
                'average': round(avg, 2),
                'total_credits': stats['total_credits'],
            })

    # 전체 평균 등급 계산 (동일 가중치)
    total_weighted_sum = sum(s['total_weighted'] for s in semester_stats.values())
    total_credits_sum = sum(s['total_credits'] for s in semester_stats.values())
    overall_average = None
    if total_credits_sum > 0:
        overall_average = round(Decimal(total_weighted_sum) / Decimal(total_credits_sum), 2)

    # 학년별 가중치 적용 전체 등급 계산
    weighted_averages = []
    weight_configs = [
        {'name': '30:30:40', 'weights': {1: 30, 2: 30, 3: 40}},
        {'name': '20:40:40', 'weights': {1: 20, 2: 40, 3: 40}},
        {'name': '20:30:50', 'weights': {1: 20, 2: 30, 3: 50}},
    ]

    for config in weight_configs:
        weights = config['weights']
        weighted_sum = Decimal(0)
        weight_sum = Decimal(0)

        for year in [1, 2, 3]:
            if year in year_stats and year_stats[year]['total_credits'] > 0:
                year_avg = Decimal(year_stats[year]['total_weighted']) / Decimal(year_stats[year]['total_credits'])
                weighted_sum += year_avg * Decimal(weights[year])
                weight_sum += Decimal(weights[year])

        if weight_sum > 0:
            weighted_avg = round(weighted_sum / weight_sum, 2)
            weighted_averages.append({
                'name': config['name'],
                'average': weighted_avg,
            })

    # 차트용 데이터 준비 (교과별 평균 등급)
    chart_data = []
    for semester_key in sorted(semester_category_grades.keys()):
        year, sem = semester_key.split('-')
        semester_data = {
            'label': f"{year}학년 {sem}학기",
            'categories': {}
        }
        for category, stats in semester_category_grades[semester_key].items():
            if stats['total_credits'] > 0:
                avg_grade = round(float(stats['total_weighted']) / float(stats['total_credits']), 2)
                semester_data['categories'][category] = {
                    'average': avg_grade,
                    'total_credits': stats['total_credits'],
                }
        chart_data.append(semester_data)

    # 교과 조합별 평균 분석
    category_combinations = [
        {'name': '국수영과', 'categories': ['국어', '수학', '영어', '과학']},
        {'name': '국수영사', 'categories': ['국어', '수학', '영어', '사회']},
        {'name': '국수영사과', 'categories': ['국어', '수학', '영어', '사회', '과학']},
    ]

    combination_averages = []
    for combo in category_combinations:
        total_weighted = 0
        total_credits = 0
        missing_categories = []

        for cat in combo['categories']:
            if cat in category_stats and category_stats[cat]['total_credits'] > 0:
                total_weighted += category_stats[cat]['total_weighted']
                total_credits += category_stats[cat]['total_credits']
            else:
                missing_categories.append(cat)

        if total_credits > 0:
            avg = round(Decimal(total_weighted) / Decimal(total_credits), 2)
            combination_averages.append({
                'name': combo['name'],
                'categories': combo['categories'],
                'average': avg,
                'total_credits': total_credits,
                'missing': missing_categories,
            })

    # 일반 내신 성적과 진로선택 성적 분리
    regular_internal_grades = [g for g in internal_grades if not g.is_elective]
    elective_grades = [g for g in internal_grades if g.is_elective]

    # 중학생 성적 요약 (원점수·평균·표준편차 기반 상위% 계산)
    is_middle = student.grade in ['K7', 'K8', 'K9']
    middle_grade_summary = []
    if is_middle:
        import math
        def _norm_cdf(x):
            return (1 + math.erf(x / math.sqrt(2))) / 2
        for g in sorted(internal_grades, key=lambda g: (g.year, g.semester, g.subject.subject_code if g.subject else '')):
            entry = {
                'grade': g,
                'upper_pct': None,
                'est_rank': None,
            }
            if g.score is not None and g.subject_average is not None and g.subject_stddev and float(g.subject_stddev) > 0:
                z = (float(g.score) - float(g.subject_average)) / float(g.subject_stddev)
                upper_pct = round((1 - _norm_cdf(z)) * 100, 1)
                entry['upper_pct'] = upper_pct
                if g.enrolled_count:
                    entry['est_rank'] = max(1, round(upper_pct / 100 * g.enrolled_count))
            middle_grade_summary.append(entry)

    context = {
        'student': student,
        'unpaid_sales': unpaid_sales,
        'paid_sales': paid_sales,
        'total_unpaid_books': total_unpaid_books,
        'total_paid_books': total_paid_books,
        'active_enrollments': active_enrollments,
        'current_month_tuition': current_month_tuition,
        'current_me_items': current_me_items,
        'past_unpaid_tuition': past_unpaid_tuition,
        'past_unpaid_me_items': past_unpaid_me_items,
        'special_enrollment_items': special_enrollment_items,
        'total_tuition_paid': total_tuition_paid,
        'tuition_payments': tuition_payments,
        'withdrawal_refunds': withdrawal_refunds,
        'combined_tuition_history': combined_tuition_history,
        'total_tuition_refund': total_tuition_refund,
        'total_tuition_net': total_tuition_net,
        'internal_grades': regular_internal_grades,
        'elective_grades': elective_grades,
        'mock_grades': mock_grades,
        'semester_averages': semester_averages,
        'overall_average': overall_average,
        'weighted_averages': weighted_averages,
        'combination_averages': combination_averages,
        'chart_data': json.dumps(chart_data, ensure_ascii=False),
        'today': today,
        'paid_this_month_lesson_ids': paid_this_month_lesson_ids,
        'is_middle': is_middle,
        'middle_grade_summary': middle_grade_summary,
        'cert_years': sorted(
            TuitionPayment.objects.filter(enrollment__student=student)
            .values_list('year', flat=True).distinct(),
            reverse=True,
        ),
    }
    return render(request, 'students/student_detail.html', context)


@login_required
def student_tuition_pay(request, pk, enroll_pk):
    from classes.models import Enrollment, TuitionPayment
    student = get_object_or_404(Student, pk=pk)
    enrollment = get_object_or_404(Enrollment, pk=enroll_pk, student=student)

    if request.method == 'POST':
        try:
            year = int(request.POST.get('year'))
            month = int(request.POST.get('month'))
            amount = int(request.POST.get('amount'))
            payment_date = request.POST.get('payment_date')
            payment_method = request.POST.get('payment_method')

            TuitionPayment.objects.create(
                enrollment=enrollment,
                year=year,
                month=month,
                amount=amount,
                payment_date=payment_date,
                payment_method=payment_method,
            )
            messages.success(request, f'{year}년 {month}월 수강료({enrollment.lesson.name}) 납부 처리되었습니다.')
        except (ValueError, TypeError):
            messages.error(request, '입력값이 올바르지 않습니다.')
    return redirect('students:student_detail', pk=pk)


@login_required
def tuition_payment_edit(request, pk):
    """수강료 납부 내역 수정 (연도/월/금액/납부방법/납부일)"""
    from classes.models import TuitionPayment
    payment = get_object_or_404(TuitionPayment, pk=pk)
    student = payment.enrollment.student

    if request.method == 'POST':
        try:
            payment.year = int(request.POST.get('year'))
            payment.month = int(request.POST.get('month'))
            payment.amount = int(request.POST.get('amount'))
            payment.payment_date = request.POST.get('payment_date')
            payment.payment_method = request.POST.get('payment_method')
            payment.save()
            messages.success(request, '수강료 납부 내역이 수정되었습니다.')
        except (ValueError, TypeError) as e:
            messages.error(request, f'입력값이 올바르지 않습니다: {e}')

    return redirect('students:student_detail', pk=student.pk)


@login_required
def tuition_payment_delete(request, pk):
    """수강료 납부 내역 삭제"""
    from classes.models import TuitionPayment
    payment = get_object_or_404(TuitionPayment, pk=pk)
    student = payment.enrollment.student

    if request.method == 'POST':
        payment.delete()
        messages.success(request, '수강료 납부 내역이 삭제되었습니다.')

    return redirect('students:student_detail', pk=student.pk)


@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, '학생 정보가 성공적으로 수정되었습니다.')
            return redirect('students:student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'students/student_form.html', {'form': form})


@login_required
def student_import(request):
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = form.cleaned_data['file']
                file_ext = file.name.split('.')[-1].lower()

                # 헬퍼 함수: 빈값 체크
                def is_empty(value):
                    return value is None or str(value).strip() == ''

                # 날짜 필드 처리 함수
                def parse_date(value):
                    if is_empty(value):
                        return None
                    elif isinstance(value, (datetime.date, datetime.datetime)):
                        return value.date() if isinstance(value, datetime.datetime) else value
                    elif isinstance(value, str):
                        for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d'):
                            try:
                                return datetime.datetime.strptime(value.strip(), fmt).date()
                            except ValueError:
                                continue
                        return None  # 지원되지 않는 형식일 경우
                    else:
                        return None

                # 파일 읽기
                rows = []
                if file_ext == 'csv':
                    content = file.read()
                    try:
                        decoded = content.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        decoded = content.decode('cp949')
                    reader = csv.DictReader(io.StringIO(decoded))
                    rows = list(reader)
                else:
                    wb = load_workbook(file, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    for excel_row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, excel_row)))
                    wb.close()

                # 중복 체크를 위한 카운터 초기화
                new_count = 0
                duplicate_count = 0

                # 각 행 처리
                for row in rows:
                    school_name = str(row.get('school') or '').strip()
                    if school_name:
                        school, created = School.objects.get_or_create(name=school_name)
                    else:
                        school = None

                    first_class_date = parse_date(row.get('first_class_date'))
                    quit_date = parse_date(row.get('quit_date'))

                    student_data = {
                        'school': school,
                        'grade': row.get('grade'),
                        'phone_number': row.get('phone_number'),
                        'email': row.get('email'),
                        'gender': row.get('gender'),
                        'parent_phone': row.get('parent_phone'),
                        'receipt_number': row.get('receipt_number'),
                        'first_class_date': first_class_date,
                        'quit_date': quit_date,
                        'etc': row.get('etc'),
                        # 필요한 다른 필드 추가
                    }

                    student, created = Student.objects.update_or_create(
                        name=row['name'],
                        defaults=student_data
                    )
                    if created:
                        new_count += 1
                    else:
                        duplicate_count += 1

                messages.success(request, f"{new_count}명 학생이 새로 추가되었고, {duplicate_count}명 학생은 이미 존재하여 업데이트되었습니다.")
                return redirect('students:student_list')
            except Exception as e:
                # 에러 로그는 서버에 기록하고 사용자에게는 일반 메시지만 표시
                logger.error(f"Student import error: {str(e)}")
                messages.error(request, "파일 처리 중 오류가 발생했습니다. 파일 형식을 확인해주세요.")
        else:
            messages.error(request, "폼이 유효하지 않습니다.")
    else:
        form = StudentImportForm()
    return render(request, 'students/student_import.html', {'form': form})


@login_required
def student_export(request):
    students = Student.objects.all()

    # openpyxl로 Excel 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = '학생 목록'

    # 헤더 작성
    headers = ['name', 'school', 'grade', 'phone_number', 'email', 'gender',
               'parent_phone', 'receipt_number', 'interview_date', 'interview_score',
               'interview_info', 'first_class_date', 'quit_date', 'etc']
    ws.append(headers)

    # 데이터 작성
    for student in students:
        row = [
            student.name,
            student.school.name if student.school else '',
            student.grade,
            student.phone_number,
            student.email,
            student.get_gender_display(),
            student.parent_phone,
            student.receipt_number,
            student.interview_date.strftime('%Y-%m-%d') if student.interview_date else '',
            student.interview_score,
            student.interview_info,
            student.first_class_date.strftime('%Y-%m-%d') if student.first_class_date else '',
            student.quit_date.strftime('%Y-%m-%d') if student.quit_date else '',
            student.etc,
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
    wb.save(response)

    return response


@login_required
def student_import_sample(request):
    """학생 데이터 가져오기용 샘플 파일 다운로드"""
    file_format = request.GET.get('format', 'xlsx')

    # 샘플 데이터
    headers = ['name', 'school', 'grade', 'phone_number', 'email', 'gender',
               'parent_phone', 'receipt_number', 'interview_date', 'interview_score',
               'interview_info', 'first_class_date', 'quit_date', 'etc']

    sample_rows = [
        ['홍길동', '서울중학교', 'K7', '010-1234-5678', 'hong@example.com', 'M',
         '010-1111-2222', '010-1111-2222', '2024-01-15', 8, '성실한 학생', '2024-02-01', '', '특이사항 없음'],
        ['김철수', '한국고등학교', 'K10', '010-9876-5432', 'kim@example.com', 'F',
         '010-3333-4444', '010-3333-4444', '2024-02-20', 7, '수학에 관심이 많음', '2024-03-01', '', ''],
    ]

    if file_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="student_import_sample.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '학생 샘플'
        ws.append(headers)
        for row in sample_rows:
            ws.append(row)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="student_import_sample.xlsx"'
        wb.save(response)

    return response

@login_required
def student_files(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        form = StudentFileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            description = form.cleaned_data.get('description', '')

            try:
                student_file = StudentFile(
                    student=student,
                    file=file,
                    file_name=file.name,
                    description=description
                )
                student_file.save()
                messages.success(request, '파일이 성공적으로 업로드되었습니다.')
            except Exception as e:
                logger.error(f"File upload error for student {pk}: {str(e)}")
                messages.error(request, '파일 업로드 중 오류가 발생했습니다.')

            return redirect('students:student_files', pk=pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, error)
    else:
        form = StudentFileUploadForm()

    files = student.files.all()
    context = {
        'student': student,
        'files': files,
        'form': form,
    }
    return render(request, 'students/student_files.html', context)

@login_required
def delete_student_file(request, file_id):
    file = get_object_or_404(StudentFile, id=file_id)
    student_pk = file.student.pk

    if request.method == 'POST':
        file.file.delete()  # 실제 파일 삭제
        file.delete()       # DB 레코드 삭제
        messages.success(request, '파일이 삭제되었습니다.')

    return redirect('students:student_files', pk=student_pk)


@login_required
def bulk_sms_send(request):
    """여러 학생/학부모에게 일괄 문자 발송 - 독립 페이지"""

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        target = request.POST.get('target')
        message = request.POST.get('message')

        if not student_ids:
            messages.error(request, '발송 대상 학생을 선택해주세요.')
            return redirect('students:bulk_sms_send')

        if not message:
            messages.error(request, '메시지를 입력해주세요.')
            return redirect('students:bulk_sms_send')

        # 선택된 학생들 조회
        students = Student.objects.filter(pk__in=student_ids, is_active=True)

        success_count = 0
        fail_messages = []

        for student in students:
            # 학생에게 발송
            if target in ['student', 'both'] and student.phone_number:
                phone = student.phone_number.replace('-', '').strip()
                is_success, msg = send_sms(phone, message)
                if is_success:
                    success_count += 1
                else:
                    fail_messages.append(f"{student.name}(학생): {msg}")

            # 학부모에게 발송
            if target in ['parent', 'both'] and student.parent_phone:
                phone = student.parent_phone.replace('-', '').strip()
                is_success, msg = send_sms(phone, message)
                if is_success:
                    success_count += 1
                else:
                    fail_messages.append(f"{student.name}(학부모): {msg}")

        # 결과 메시지 처리
        if success_count > 0:
            messages.success(request, f"{success_count}건의 문자를 발송했습니다.")

        if fail_messages:
            for f_msg in fail_messages[:10]:  # 최대 10개까지만 표시
                messages.error(request, f_msg)
            if len(fail_messages) > 10:
                messages.error(request, f"외 {len(fail_messages) - 10}건의 발송 실패")

        return redirect('students:bulk_sms_send')

    # GET 요청: 모든 재원 중인 학생 목록 표시
    students = Student.objects.filter(is_active=True).select_related('school').order_by('grade', 'name')

    # 학년별 그룹화
    from collections import defaultdict
    grade_order = ['K5', 'K6', 'K7', 'K8', 'K9', 'K10', 'K11', 'K12']
    grouped = defaultdict(list)

    for student in students:
        grade = student.grade if student.grade else '미지정'
        grouped[grade].append(student)

    # 학년 순서대로 정렬
    sorted_grouped = {}
    for grade in grade_order:
        if grade in grouped:
            sorted_grouped[grade] = grouped[grade]
    # 미지정 추가
    for grade, student_list in grouped.items():
        if grade not in grade_order:
            sorted_grouped[grade] = student_list

    return render(request, 'students/bulk_sms_page.html', {
        'grouped_students': sorted_grouped,
    })


@login_required
def student_send_email(request, pk):
    """학생에게 이메일 발송"""
    from django.core.mail import EmailMessage
    from django.conf import settings
    from .forms import StudentEmailForm

    student = get_object_or_404(Student, pk=pk)

    if not student.email:
        messages.error(request, '해당 학생의 이메일 주소가 등록되어 있지 않습니다.')
        return redirect('students:student_detail', pk=pk)

    if request.method == 'POST':
        form = StudentEmailForm(request.POST, request.FILES)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']

            # 발신자 이메일 설정
            from_email = settings.DEFAULT_FROM_EMAIL if request.user.username == 'admin' else settings.EMAIL_HOST_USER

            try:
                # EmailMessage 객체 생성
                email = EmailMessage(
                    subject=subject,
                    body=message,
                    from_email=from_email,
                    to=[student.email],
                )

                # 첨부파일 처리
                files = request.FILES.getlist('attachments')
                for file in files:
                    email.attach(file.name, file.read(), file.content_type)

                # 이메일 전송
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f'Sending email to {student.email} from {from_email}')

                result = email.send(fail_silently=False)

                logger.info(f'Email send result: {result}')

                messages.success(request, f'{student.name} 학생에게 이메일을 성공적으로 발송했습니다.')
                return redirect('students:student_detail', pk=pk)
            except Exception as e:
                logger.error(f'Email sending failed for student {pk}: {str(e)}', exc_info=True)
                messages.error(request, '이메일 발송 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.')
    else:
        form = StudentEmailForm()

    context = {
        'form': form,
        'student': student,
    }
    return render(request, 'students/student_email_form.html', context)


@login_required
def student_send_sms(request, pk):
    """학생에게 문자 발송"""
    from .forms import StudentSMSForm
    import requests

    student = get_object_or_404(Student, pk=pk)

    if not student.phone_number:
        messages.error(request, '해당 학생의 전화번호가 등록되어 있지 않습니다.')
        return redirect('students:student_detail', pk=pk)

    if request.method == 'POST':
        form = StudentSMSForm(request.POST)
        if form.is_valid():
            message = form.cleaned_data['message']

            try:
                # Aligo SMS API 설정
                from django.conf import settings

                sms_url = 'https://apis.aligo.in/send/'
                sms_data = {
                    'key': settings.SMS_API_KEY,
                    'user_id': settings.SMS_USER_ID,
                    'sender': settings.SMS_SENDER_NUMBER,
                    'receiver': student.phone_number,
                    'msg': message,
                    'msg_type': 'LMS' if len(message.encode('euc-kr')) > 90 else 'SMS',
                    'title': '엠클래스' if len(message.encode('euc-kr')) > 90 else '',
                }

                response = requests.post(sms_url, data=sms_data)
                result = response.json()

                if result.get('result_code') == '1':
                    messages.success(request, f'{student.name} 학생에게 문자를 성공적으로 발송했습니다.')
                else:
                    messages.error(request, f'문자 발송 실패: {result.get("message", "알 수 없는 오류")}')

                return redirect('students:student_detail', pk=pk)
            except Exception as e:
                logger.error(f'SMS sending failed for student {pk}: {str(e)}', exc_info=True)
                messages.error(request, '문자 발송 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.')
    else:
        form = StudentSMSForm()

    context = {
        'form': form,
        'student': student,
    }
    return render(request, 'students/student_sms_form.html', context)


@login_required
def student_send_sms_parent(request, pk):
    """부모님에게 문자 발송"""
    from .forms import StudentSMSForm
    import requests

    student = get_object_or_404(Student, pk=pk)

    if not student.parent_phone:
        messages.error(request, '해당 학생의 부모님 전화번호가 등록되어 있지 않습니다.')
        return redirect('students:student_detail', pk=pk)

    if request.method == 'POST':
        form = StudentSMSForm(request.POST)
        if form.is_valid():
            message = form.cleaned_data['message']

            try:
                # Aligo SMS API 설정
                from django.conf import settings

                sms_url = 'https://apis.aligo.in/send/'
                sms_data = {
                    'key': settings.SMS_API_KEY,
                    'user_id': settings.SMS_USER_ID,
                    'sender': settings.SMS_SENDER_NUMBER,
                    'receiver': student.parent_phone,
                    'msg': message,
                    'msg_type': 'LMS' if len(message.encode('euc-kr')) > 90 else 'SMS',
                    'title': '엠클래스' if len(message.encode('euc-kr')) > 90 else '',
                }

                response = requests.post(sms_url, data=sms_data)
                result = response.json()

                if result.get('result_code') == '1':
                    messages.success(request, f'{student.name} 학생 부모님에게 문자를 성공적으로 발송했습니다.')
                else:
                    messages.error(request, f'문자 발송 실패: {result.get("message", "알 수 없는 오류")}')

                return redirect('students:student_detail', pk=pk)
            except Exception as e:
                logger.error(f'SMS sending to parent failed for student {pk}: {str(e)}', exc_info=True)
                messages.error(request, '문자 발송 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.')
    else:
        form = StudentSMSForm()

    context = {
        'form': form,
        'student': student,
        'is_parent': True,  # 부모님에게 보내는 것임을 표시
    }
    return render(request, 'students/student_sms_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def grade_promotion_confirm(request):
    """학년 일괄 증가 확인 페이지 (관리자 전용)"""
    # 현재 학년별 학생 수 집계
    from django.db.models import Count

    grade_stats = Student.objects.filter(
        is_active=True
    ).values('grade').annotate(
        count=Count('id')
    ).order_by('grade')

    # 학년 증가 후 예상 결과 계산
    grade_mapping = {
        'K5': 'K6', 'K6': 'K7', 'K7': 'K8', 'K8': 'K9',
        'K9': 'K10', 'K10': 'K11', 'K11': 'K12', 'K12': '졸업'
    }

    changes = []
    for stat in grade_stats:
        current_grade = stat['grade']
        count = stat['count']
        new_grade = grade_mapping.get(current_grade, current_grade)

        # 해당 학년의 학생 이름 목록 조회
        students = Student.objects.filter(
            is_active=True,
            grade=current_grade
        ).order_by('name')

        changes.append({
            'current': current_grade,
            'new': new_grade,
            'count': count,
            'students': students
        })

    context = {
        'changes': changes,
        'total_students': sum(stat['count'] for stat in grade_stats),
    }
    return render(request, 'students/grade_promotion_confirm.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def grade_promotion_execute(request):
    """학년 일괄 증가 실행 (관리자 전용)"""
    if request.method != 'POST':
        messages.error(request, '잘못된 접근입니다.')
        return redirect('students:student_list')

    # 학년 매핑
    grade_mapping = {
        'K5': 'K6', 'K6': 'K7', 'K7': 'K8', 'K8': 'K9',
        'K9': 'K10', 'K10': 'K11', 'K11': 'K12', 'K12': '졸업'
    }

    # 활성 학생들만 대상
    active_students = Student.objects.filter(is_active=True)

    updated_count = 0
    graduated_count = 0

    for student in active_students:
        if student.grade in grade_mapping:
            new_grade = grade_mapping[student.grade]
            if new_grade == '졸업':
                # K12 학생은 졸업 처리
                student.grade = '졸업'
                student.is_active = False
                student.quit_date = datetime.date.today()
                graduated_count += 1
            else:
                student.grade = new_grade
                updated_count += 1
            student.save()

    messages.success(
        request,
        f'학년 증가가 완료되었습니다. (진급: {updated_count}명, 졸업: {graduated_count}명)'
    )
    return redirect('students:student_list')


@login_required
def student_quit(request, pk):
    """학생 퇴원 처리 (환불 계산 포함)"""
    import datetime as _dt
    from classes.models import Enrollment
    from classes.utils import calculate_refund

    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        from classes.models import WithdrawalRefund

        quit_date_str = request.POST.get('quit_date', '')
        try:
            quit_date = _dt.date.fromisoformat(quit_date_str)
        except ValueError:
            quit_date = _dt.date.today()

        student.quit_date = quit_date
        student.is_active = False
        student.save()

        # 활성 수강별 환불 계산 → 기록 저장 + 수강 종료
        active_enrollments = Enrollment.objects.filter(
            student=student, is_active=True
        ).select_related('lesson')

        total_refund = 0
        for enroll in active_enrollments:
            info = calculate_refund(enroll, quit_date)
            # 환불 금액이 있는 수업만 저장 (수업 일정 없거나 환불율 0%인 경우 제외)
            if info['total_days'] > 0:
                WithdrawalRefund.objects.create(
                    enrollment=enroll,
                    student=student,
                    quit_date=quit_date,
                    year=quit_date.year,
                    month=quit_date.month,
                    tuition=info['tuition'],
                    total_days=info['total_days'],
                    passed_days=info['passed_days'],
                    refund_rate=info['refund_rate'],
                    refund_amount=info['refund_amount'],
                )
                total_refund += info['refund_amount']

            enroll.is_active = False
            enroll.end_date = quit_date
            enroll.save()

        refund_msg = f' 환불 예정액: {total_refund:,}원' if total_refund > 0 else ''
        messages.success(
            request,
            f'{student.name} 학생이 퇴원 처리되었습니다. (퇴원일: {quit_date}{refund_msg})'
        )
        return redirect('students:student_detail', pk=pk)

    # GET: 환불 금액 미리보기
    today = _dt.date.today()
    active_enrollments = Enrollment.objects.filter(
        student=student, is_active=True
    ).select_related('lesson')

    quit_date_str = request.GET.get('quit_date', today.isoformat())
    try:
        quit_date = _dt.date.fromisoformat(quit_date_str)
    except ValueError:
        quit_date = today

    refund_items = []
    total_refund = 0
    for enroll in active_enrollments:
        info = calculate_refund(enroll, quit_date)
        info['enrollment'] = enroll
        info['lesson_name'] = enroll.lesson.name
        refund_items.append(info)
        total_refund += info['refund_amount']

    return render(request, 'students/student_quit_confirm.html', {
        'student': student,
        'quit_date': quit_date,
        'refund_items': refund_items,
        'total_refund': total_refund,
    })


@login_required
def student_readmit(request, pk):
    """학생 재입원 처리"""
    student = get_object_or_404(Student, pk=pk)

    if request.method == 'POST':
        student.quit_date = None
        student.is_active = True
        student.save()
        messages.success(request, f'{student.name} 학생이 재입원 처리되었습니다.')
        return redirect('students:student_detail', pk=pk)

    return render(request, 'students/student_readmit_confirm.html', {'student': student})


def parent_lookup(request):
    """학부모 교재 결제 내역 조회 (로그인 불필요)"""
    import datetime as _dt
    from django.db.models import Q as _Q
    from bookstore.models import BookSale
    from classes.models import MonthlyEnrollment, TuitionPayment

    student = None
    unpaid_sales = []
    paid_sales = []
    total_unpaid = 0
    total_paid = 0
    error_message = None
    this_month_tuition = 0
    this_month_book_total = 0
    # 성적 분석용 변수 초기화
    grade_internal_grades = []
    grade_elective_grades = []
    grade_mock_grades = []
    grade_semester_averages = []
    grade_overall_average = None
    grade_weighted_averages = []
    grade_combination_averages = []
    grade_chart_data = '[]'
    grade_semester_averages_json = '[]'
    grade_is_middle = False
    grade_middle_grade_summary = []
    grade_groups = []
    this_month_total = 0
    current_me_items = []
    month_book_sales = []
    past_unpaid_mes = []
    current_unpaid_mes = []
    past_unpaid_tuition = 0
    total_unpaid_tuition = 0
    paid_tuition_payments = []
    total_paid_tuition = 0
    total_unpaid_all = 0
    total_paid_all = 0

    today = _dt.date.today()
    this_year = today.year
    this_month = today.month
    month_start = today.replace(day=1)

    # 세션에서 학생 정보 복원 시도
    student_name = request.POST.get('student_name', '').strip()
    student_id = request.POST.get('student_id', '').strip()

    # POST 요청이 아니면 세션에서 복원
    if request.method != 'POST' and 'parent_student_id' in request.session:
        student_id = request.session.get('parent_student_id', '')
        student_name = request.session.get('parent_student_name', '')

    if student_name and student_id:
        try:
            student = Student.objects.get(name=student_name, student_id=student_id)

            # 세션에 학생 정보 저장
            request.session['parent_student_id'] = student_id
            request.session['parent_student_name'] = student_name

            # 결제 완료 내역
            paid_sales = BookSale.objects.filter(
                student=student, is_paid=True
            ).select_related('book').order_by('-payment_date')
            total_paid = sum(sale.get_total_price() for sale in paid_sales)

            # 전체 수강료 납부 기록 (학생)
            paid_quads = set(
                TuitionPayment.objects.filter(
                    enrollment__student=student,
                ).values_list('enrollment__lesson_id', 'year', 'month')
            )

            # 이번 달 ME (cancelled 제외)
            current_mes = list(
                MonthlyEnrollment.objects.filter(
                    student=student,
                    year=this_year,
                    month=this_month,
                ).exclude(status='cancelled').select_related('lesson').order_by('lesson__name')
            )
            current_me_items = [
                {'me': me, 'is_paid': (me.lesson_id, me.year, me.month) in paid_quads}
                for me in current_mes
            ]
            this_month_tuition = sum(item['me'].adjusted_tuition for item in current_me_items)

            # 미결제 교재 전체 → 청구 파일과 동일하게 날짜 무관 is_paid=False 전체
            month_book_sales = list(BookSale.objects.filter(
                student=student,
                is_paid=False,
            ).select_related('book').order_by('-sale_date'))
            this_month_book_total = sum(s.get_total_price() for s in month_book_sales)

            # 교재는 청구에 모두 포함되므로 미결제 내역에서 제외
            unpaid_sales = []
            total_unpaid = 0

            this_month_total = this_month_tuition + this_month_book_total

            # 과거 미납 ME (이번 달 이전, 납부 기록 없는 것)
            past_all_mes = list(
                MonthlyEnrollment.objects.filter(
                    student=student,
                ).exclude(status='cancelled').filter(
                    _Q(year__lt=this_year) | _Q(year=this_year, month__lt=this_month)
                ).select_related('lesson').order_by('-year', '-month', 'lesson__name')
            )
            past_unpaid_mes = [
                me for me in past_all_mes
                if (me.lesson_id, me.year, me.month) not in paid_quads
            ]
            past_unpaid_tuition = sum(me.adjusted_tuition for me in past_unpaid_mes)

            # 이번 달 미납 ME
            current_unpaid_mes = [item['me'] for item in current_me_items if not item['is_paid']]
            this_month_unpaid_tuition = sum(me.adjusted_tuition for me in current_unpaid_mes)

            total_unpaid_tuition = past_unpaid_tuition

            # 납부 완료 수강료: 전체 납부 이력 (최신순)
            paid_tuition_payments = list(
                TuitionPayment.objects.filter(
                    enrollment__student=student,
                ).select_related('enrollment__lesson').order_by('-year', '-month')
            )
            total_paid_tuition = sum(p.amount for p in paid_tuition_payments)

            total_unpaid_all = total_unpaid + total_unpaid_tuition
            total_paid_all = total_paid + total_paid_tuition

            # ── 성적 분석 데이터 ────────────────────────────────
            from grades.models import Grade as _Grade
            from collections import defaultdict as _defaultdict
            from decimal import Decimal as _Decimal
            import json as _json
            import math as _math

            _internal_grades = _Grade.objects.filter(
                student=student, grade_type='internal'
            ).select_related('subject').order_by('-year', '-semester', 'subject__subject_code')

            _mock_grades = _Grade.objects.filter(
                student=student, grade_type='mock'
            ).select_related('subject').order_by('-exam_year', '-exam_month', 'subject__subject_code')

            _semester_stats = _defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})
            _year_stats = _defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})
            _semester_category_grades = _defaultdict(lambda: _defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0}))
            _category_stats = _defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})

            for _g in _internal_grades:
                if _g.grade_rank is None or _g.credits is None:
                    continue
                if not _g.is_elective:
                    _key = f"{_g.year}-{_g.semester}"
                    _cat = _g.curriculum or '기타'
                    _semester_category_grades[_key][_cat]['total_weighted'] += _g.grade_rank * _g.credits
                    _semester_category_grades[_key][_cat]['total_credits'] += _g.credits
                    _category_stats[_cat]['total_weighted'] += _g.grade_rank * _g.credits
                    _category_stats[_cat]['total_credits'] += _g.credits
                if _g.is_elective:
                    continue
                _sem_key = (_g.year, _g.semester)
                _semester_stats[_sem_key]['total_weighted'] += _g.grade_rank * _g.credits
                _semester_stats[_sem_key]['total_credits'] += _g.credits
                _year_stats[_g.year]['total_weighted'] += _g.grade_rank * _g.credits
                _year_stats[_g.year]['total_credits'] += _g.credits

            _semester_averages = []
            for (_yr, _sem), _stats in sorted(_semester_stats.items()):
                if _stats['total_credits'] > 0:
                    _avg = _Decimal(_stats['total_weighted']) / _Decimal(_stats['total_credits'])
                    _semester_averages.append({
                        'year': _yr, 'semester': _sem,
                        'average': round(_avg, 2),
                        'total_credits': _stats['total_credits'],
                    })

            _total_wsum = sum(s['total_weighted'] for s in _semester_stats.values())
            _total_csum = sum(s['total_credits'] for s in _semester_stats.values())
            _overall_average = None
            if _total_csum > 0:
                _overall_average = round(_Decimal(_total_wsum) / _Decimal(_total_csum), 2)

            _weight_configs = [
                {'name': '30:30:40', 'weights': {1: 30, 2: 30, 3: 40}},
                {'name': '20:40:40', 'weights': {1: 20, 2: 40, 3: 40}},
                {'name': '20:30:50', 'weights': {1: 20, 2: 30, 3: 50}},
            ]
            _weighted_averages = []
            for _config in _weight_configs:
                _wts = _config['weights']
                _wsum = _Decimal(0)
                _wt_sum = _Decimal(0)
                for _yr in [1, 2, 3]:
                    if _yr in _year_stats and _year_stats[_yr]['total_credits'] > 0:
                        _yr_avg = _Decimal(_year_stats[_yr]['total_weighted']) / _Decimal(_year_stats[_yr]['total_credits'])
                        _wsum += _yr_avg * _Decimal(_wts[_yr])
                        _wt_sum += _Decimal(_wts[_yr])
                if _wt_sum > 0:
                    _weighted_averages.append({'name': _config['name'], 'average': round(_wsum / _wt_sum, 2)})

            _chart_data = []
            for _sem_key in sorted(_semester_category_grades.keys()):
                _yr, _sm = _sem_key.split('-')
                _sem_data = {'label': f"{_yr}학년 {_sm}학기", 'categories': {}}
                for _cat, _stats in _semester_category_grades[_sem_key].items():
                    if _stats['total_credits'] > 0:
                        _avg_g = round(float(_stats['total_weighted']) / float(_stats['total_credits']), 2)
                        _sem_data['categories'][_cat] = {'average': _avg_g, 'total_credits': _stats['total_credits']}
                _chart_data.append(_sem_data)

            _category_combinations = [
                {'name': '국수영과', 'categories': ['국어', '수학', '영어', '과학']},
                {'name': '국수영사', 'categories': ['국어', '수학', '영어', '사회']},
                {'name': '국수영사과', 'categories': ['국어', '수학', '영어', '사회', '과학']},
            ]
            _combination_averages = []
            for _combo in _category_combinations:
                _tw = 0; _tc = 0; _miss = []
                for _cat in _combo['categories']:
                    if _cat in _category_stats and _category_stats[_cat]['total_credits'] > 0:
                        _tw += _category_stats[_cat]['total_weighted']
                        _tc += _category_stats[_cat]['total_credits']
                    else:
                        _miss.append(_cat)
                if _tc > 0:
                    _combination_averages.append({
                        'name': _combo['name'], 'categories': _combo['categories'],
                        'average': round(_Decimal(_tw) / _Decimal(_tc), 2),
                        'total_credits': _tc, 'missing': _miss,
                    })

            _regular_internal = [g for g in _internal_grades if not g.is_elective]
            _elective = [g for g in _internal_grades if g.is_elective]

            _is_middle = student.grade in ['K7', 'K8', 'K9']
            _middle_summary = []
            if _is_middle:
                def _norm_cdf(x):
                    return (1 + _math.erf(x / _math.sqrt(2))) / 2
                for _g in sorted(_internal_grades, key=lambda g: (g.year, g.semester, g.subject.subject_code if g.subject else '')):
                    _entry = {'grade': _g, 'upper_pct': None, 'est_rank': None}
                    if _g.score is not None and _g.subject_average is not None and _g.subject_stddev and float(_g.subject_stddev) > 0:
                        _z = (float(_g.score) - float(_g.subject_average)) / float(_g.subject_stddev)
                        _up = round((1 - _norm_cdf(_z)) * 100, 1)
                        _entry['upper_pct'] = _up
                        if _g.enrolled_count:
                            _entry['est_rank'] = max(1, round(_up / 100 * _g.enrolled_count))
                    _middle_summary.append(_entry)

            # 학기별 그룹 (성적 테이블용)
            _grade_groups_dict = _defaultdict(list)
            for _g in _Grade.objects.filter(
                student=student, grade_type='internal'
            ).select_related('subject').order_by('year', 'semester', 'subject__subject_code'):
                _grade_groups_dict[(_g.year, _g.semester)].append(_g)
            _grade_groups = [
                {'year': k[0], 'semester': k[1], 'grades': v}
                for k, v in sorted(_grade_groups_dict.items(), reverse=True)
            ]

            grade_internal_grades = _regular_internal
            grade_elective_grades = _elective
            grade_mock_grades = list(_mock_grades)
            grade_semester_averages = _semester_averages
            grade_overall_average = _overall_average
            grade_weighted_averages = _weighted_averages
            grade_combination_averages = _combination_averages
            grade_chart_data = _json.dumps(_chart_data, ensure_ascii=False)
            grade_semester_averages_json = _json.dumps(
                [{'year': s['year'], 'semester': s['semester'], 'average': float(s['average'])} for s in _semester_averages],
                ensure_ascii=False
            )
            grade_is_middle = _is_middle
            grade_middle_grade_summary = _middle_summary
            grade_groups = _grade_groups

        except Student.DoesNotExist:
            error_message = '학생 정보를 찾을 수 없습니다. 이름과 고유번호를 확인해 주세요.'
    elif request.method == 'POST':
        error_message = '학생 이름과 고유번호를 모두 입력해 주세요.'

    context = {
        'student': student,
        'unpaid_sales': unpaid_sales,
        'paid_sales': paid_sales,
        'total_unpaid': total_unpaid,
        'total_paid': total_paid,
        'error_message': error_message,
        'bank_account': '신한은행 110-247-214359 장동민(엠클래스수학과학전문학원)',
        'this_month': this_month,
        'this_year': this_year,
        'current_me_items': current_me_items,
        'month_book_sales': month_book_sales,
        'this_month_tuition': this_month_tuition,
        'this_month_book_total': this_month_book_total,
        'this_month_total': this_month_total,
        'past_unpaid_mes': past_unpaid_mes,
        'past_unpaid_tuition': past_unpaid_tuition,
        'current_unpaid_mes': current_unpaid_mes,
        'total_unpaid_tuition': total_unpaid_tuition,
        'paid_tuition_payments': paid_tuition_payments,
        'total_paid_tuition': total_paid_tuition,
        'total_unpaid_all': total_unpaid_all,
        'total_paid_all': total_paid_all,
        'grade_internal_grades': grade_internal_grades,
        'grade_elective_grades': grade_elective_grades,
        'grade_mock_grades': grade_mock_grades,
        'grade_semester_averages': grade_semester_averages,
        'grade_overall_average': grade_overall_average,
        'grade_weighted_averages': grade_weighted_averages,
        'grade_combination_averages': grade_combination_averages,
        'grade_chart_data': grade_chart_data,
        'grade_semester_averages_json': grade_semester_averages_json,
        'grade_is_middle': grade_is_middle,
        'grade_middle_grade_summary': grade_middle_grade_summary,
        'grade_groups': grade_groups,
    }
    return render(request, 'students/parent_lookup.html', context)


def parent_logout(request):
    """부모님 페이지 나가기 (세션 정리)"""
    if 'parent_student_id' in request.session:
        del request.session['parent_student_id']
    if 'parent_student_name' in request.session:
        del request.session['parent_student_name']
    return redirect('index')


def parent_student_update(request, student_id):
    """부모님용 학생 정보 수정 (로그인 불필요, 제한된 필드만 수정 가능)"""
    from .forms import ParentStudentUpdateForm

    student = get_object_or_404(Student, student_id=student_id)

    if request.method == 'POST':
        form = ParentStudentUpdateForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return render(request, 'students/parent_student_update.html', {
                'student': student,
                'form': form,
                'success_message': '학생 정보가 성공적으로 수정되었습니다.'
            })
    else:
        form = ParentStudentUpdateForm(instance=student)

    return render(request, 'students/parent_student_update.html', {
        'student': student,
        'form': form,
    })


def parent_grades(request, student_pk):
    """학부모용 성적 조회 (세션 기반 인증, 로그인 불필요)"""
    import json
    from collections import defaultdict
    from grades.models import Grade
    from progress.models import LearningRecord

    # 세션 인증
    session_student_id = request.session.get('parent_student_id', '')
    session_student_name = request.session.get('parent_student_name', '')
    if not session_student_id or not session_student_name:
        return redirect('parent_lookup')

    student = get_object_or_404(Student, pk=student_pk)
    if student.student_id != session_student_id or student.name != session_student_name:
        return redirect('parent_lookup')

    # ── 퀴즈/테스트 기록 ─────────────────────────────
    TEST_TYPE_LABELS = {
        'quiz': '퀴즈', 'practice': '연습 문제',
        'booklet': '제본 교재', 'entrance_exam': '입학 시험', 'other': '기타',
    }
    raw_test_records = LearningRecord.objects.filter(
        student=student,
        record_type__in=['quiz', 'practice', 'booklet', 'entrance_exam', 'other']
    ).select_related('subject').order_by('-date', '-created_at')

    processed_test_records = []
    for rec in raw_test_records:
        quiz_detail = None
        if rec.quiz_results:
            total = rec.quiz_results.get('total', 0)
            wrong_nums = sorted(rec.quiz_results.get('wrong', []))
            correct_count = total - len(wrong_nums)
            correct_pct = round(correct_count / total * 100, 1) if total else 0
            wrong_pct = round(100 - correct_pct, 1) if total else 0
            quiz_detail = {
                'total': total,
                'wrong_nums': wrong_nums,
                'q_range': list(range(1, total + 1)),
                'correct_count': correct_count,
                'wrong_count': len(wrong_nums),
                'correct_pct': correct_pct,
                'segments_json': json.dumps([
                    {'p': correct_pct, 'c': '#22c55e'},
                    {'p': wrong_pct, 'c': '#f87171'},
                ]) if total > 0 else '[]',
            }
        processed_test_records.append({
            'rec': rec,
            'quiz_detail': quiz_detail,
            'type_label': TEST_TYPE_LABELS.get(rec.record_type, rec.record_type),
        })

    # ── 학교 시험 성적 (내신) ─────────────────────────
    GRADE_RANK_COLORS = {
        1: 'indigo', 2: 'blue', 3: 'cyan', 4: 'green',
        5: 'yellow', 6: 'orange', 7: 'red', 8: 'red', 9: 'red',
    }
    internal_grades = Grade.objects.filter(
        student=student, grade_type='internal'
    ).select_related('subject').order_by('year', 'semester', 'subject__subject_code')

    grade_groups = defaultdict(list)
    for g in internal_grades:
        grade_groups[(g.year, g.semester)].append(g)
    sorted_grade_groups = [
        {'year': k[0], 'semester': k[1], 'grades': v}
        for k, v in sorted(grade_groups.items(), reverse=True)
    ]

    # ── 진도 평가 분석 ────────────────────────────────
    ACHIEVEMENT_META = [
        {'code': 'A', 'label': '우수',   'color': 'indigo', 'hex': '#6366f1'},
        {'code': 'B', 'label': '양호',   'color': 'green',  'hex': '#22c55e'},
        {'code': 'C', 'label': '보통',   'color': 'yellow', 'hex': '#facc15'},
        {'code': 'D', 'label': '미흡',   'color': 'orange', 'hex': '#fb923c'},
        {'code': 'F', 'label': '재학습', 'color': 'red',    'hex': '#f87171'},
    ]
    textbook_records = LearningRecord.objects.filter(
        student=student,
        record_type='textbook',
        achievement__in=['A', 'B', 'C', 'D', 'F'],
    ).select_related('book_sale__book')

    book_map = {}
    for rec in textbook_records:
        if rec.book_sale_id not in book_map:
            book_map[rec.book_sale_id] = {'book_sale': rec.book_sale, 'records': []}
        book_map[rec.book_sale_id]['records'].append(rec)

    book_progress_data = []
    overall_counts = {m['code']: 0 for m in ACHIEVEMENT_META}
    for bs_id, data in book_map.items():
        counts = {m['code']: 0 for m in ACHIEVEMENT_META}
        for rec in data['records']:
            if rec.achievement in counts:
                counts[rec.achievement] += 1
                overall_counts[rec.achievement] += 1
        total = sum(counts.values())
        levels = [
            {**m, 'count': counts[m['code']],
             'percent': round(counts[m['code']] / total * 100, 1) if total else 0}
            for m in ACHIEVEMENT_META
        ]
        book_progress_data.append({
            'book_sale': data['book_sale'],
            'levels': levels,
            'total': total,
            'segments_json': json.dumps([
                {'p': l['percent'], 'c': l['hex']} for l in levels if l['percent'] > 0
            ]),
        })

    overall_total = sum(overall_counts.values())
    overall_levels = [
        {**m, 'count': overall_counts[m['code']],
         'percent': round(overall_counts[m['code']] / overall_total * 100, 1) if overall_total else 0}
        for m in ACHIEVEMENT_META
    ]
    overall_segments_json = json.dumps([
        {'p': l['percent'], 'c': l['hex']} for l in overall_levels if l['percent'] > 0
    ]) if overall_total > 0 else '[]'

    context = {
        'student': student,
        'test_records': processed_test_records,
        'grade_groups': sorted_grade_groups,
        'book_progress_data': book_progress_data,
        'overall_levels': overall_levels,
        'overall_total': overall_total,
        'overall_segments_json': overall_segments_json,
        'achievement_meta': ACHIEVEMENT_META,
        'grade_rank_colors': GRADE_RANK_COLORS,
        'is_middle': student.grade in ['K7', 'K8', 'K9'],
    }
    return render(request, 'students/parent_grades.html', context)


def _parent_auth(request, student_pk):
    """학부모 세션 인증 헬퍼. 인증된 Student 반환, 실패 시 None 반환."""
    session_id = request.session.get('parent_student_id', '')
    session_name = request.session.get('parent_student_name', '')
    if not session_id or not session_name:
        return None
    try:
        student = Student.objects.get(pk=student_pk)
        if student.student_id != session_id or student.name != session_name:
            return None
        return student
    except Student.DoesNotExist:
        return None


def parent_grade_bulk_create(request, student_pk):
    """학부모용 한 학기 내신 성적 일괄 입력 (세션 인증)"""
    from grades.models import Grade
    from subjects.models import Subject
    from decimal import Decimal, InvalidOperation

    student = _parent_auth(request, student_pk)
    if not student:
        return redirect('parent_lookup')

    is_middle = student.grade in ['K7', 'K8', 'K9']
    is_2022 = (not is_middle) and (student.curriculum_year == 2022)

    if is_middle:
        subjects = Subject.objects.filter(is_active=True, school_level='M').order_by('subject_code')
    elif is_2022:
        subjects = Subject.objects.filter(is_active=True, school_level='H', curriculum_year=2022).order_by('subject_code')
    else:
        subjects = Subject.objects.filter(is_active=True, school_level='H').exclude(curriculum_year=2022).order_by('subject_code')

    error_message = None

    if request.method == 'POST':
        year = request.POST.get('year', '').strip()
        semester = request.POST.get('semester', '').strip()
        grade_count = int(request.POST.get('grade_count', 0))

        if not year or not semester:
            error_message = '학년과 학기를 선택해 주세요.'
        else:
            created_count = 0
            for i in range(grade_count):
                subject_id = request.POST.get(f'grades[{i}][subject]', '').strip()
                score_str = request.POST.get(f'grades[{i}][score]', '').strip()

                if not subject_id or not score_str:
                    continue
                try:
                    subject = Subject.objects.get(pk=subject_id)
                    score = Decimal(score_str)

                    def _int(key):
                        v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                        return int(v) if v else None

                    def _dec(key):
                        v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                        return Decimal(v) if v else None

                    def _str(key):
                        return request.POST.get(f'grades[{i}][{key}]', '').strip() or None

                    if is_middle:
                        Grade.objects.create(
                            student=student,
                            grade_type='internal',
                            subject=subject,
                            year=int(year),
                            semester=int(semester),
                            credits=_int('credits'),
                            score=score,
                            subject_average=_dec('subject_average'),
                            subject_stddev=_dec('subject_stddev'),
                            enrolled_count=_int('enrolled_count'),
                            achievement_level=_str('achievement_level'),
                            subject_classification='general',
                            is_elective=False,
                        )
                    elif is_2022:
                        Grade.objects.create(
                            student=student,
                            grade_type='internal',
                            subject=subject,
                            year=int(year),
                            semester=int(semester),
                            credits=_int('credits'),
                            score=score,
                            subject_average=None,
                            subject_stddev=None,
                            grade_rank=_int('grade_rank'),
                            subject_classification='general',
                            is_elective=False,
                            enrolled_count=_int('enrolled_count'),
                            subject_rank=_int('subject_rank'),
                            same_rank_count=_int('same_rank_count'),
                            achievement_level=_str('achievement_level'),
                            percentile=_dec('percentile'),
                        )
                    else:
                        # 2015 교육과정
                        classification_raw = request.POST.get(f'grades[{i}][subject_classification]', 'general').strip()
                        subject_classification = classification_raw if classification_raw in ('common', 'general', 'elective', 'fusion') else 'general'
                        is_achievement = subject_classification in ('elective', 'fusion')
                        grade_rank_str = request.POST.get(f'grades[{i}][grade_rank]', '').strip()
                        grade_rank = int(grade_rank_str) if grade_rank_str and not is_achievement else None
                        Grade.objects.create(
                            student=student,
                            grade_type='internal',
                            subject=subject,
                            year=int(year),
                            semester=int(semester),
                            credits=_int('credits'),
                            score=score,
                            subject_average=_dec('subject_average'),
                            subject_stddev=_dec('subject_stddev') if not is_achievement else None,
                            grade_rank=grade_rank,
                            subject_classification=subject_classification,
                            is_elective=is_achievement,
                            achievement_level=_str('achievement_level') if is_achievement else None,
                            distribution_a=_dec('distribution_a') if is_achievement else None,
                            distribution_b=_dec('distribution_b') if is_achievement else None,
                            distribution_c=_dec('distribution_c') if is_achievement else None,
                        )
                    created_count += 1
                except Exception:
                    continue
            if created_count > 0:
                return redirect('parent_grades', student_pk=student_pk)
            error_message = '과목과 원점수를 입력해 주세요.'

    import datetime
    subjects_with_class = []
    TYPE_MAP = {'0': 'common', '1': 'general', '2': 'elective', '3': 'fusion'}
    for s in subjects:
        sc = TYPE_MAP.get(s.subject_code[2], '') if is_2022 and len(s.subject_code) == 6 else ''
        subjects_with_class.append({'id': s.pk, 'name': s.name, 'code': s.subject_code, 'category': s.category, 'classification': sc})

    context = {
        'student': student,
        'subjects_list': subjects_with_class,
        'current_year': datetime.date.today().year,
        'error_message': error_message,
        'is_2022': is_2022,
        'is_middle': is_middle,
    }
    if is_middle:
        template = 'students/parent_grade_bulk_middle.html'
    elif is_2022:
        template = 'students/parent_grade_bulk_2022.html'
    else:
        template = 'students/parent_grade_bulk.html'
    return render(request, template, context)


def parent_mock_grade_create(request, student_pk):
    """학부모용 2022 내신 성적 일괄 입력 (세션 인증)"""
    from grades.models import Grade
    from subjects.models import Subject
    from decimal import Decimal

    student = _parent_auth(request, student_pk)
    if not student:
        return redirect('parent_lookup')

    subjects = Subject.objects.filter(is_active=True, curriculum_year=2022).order_by('subject_code')
    TYPE_MAP = {'0': 'common', '1': 'general', '2': 'elective', '3': 'fusion'}
    subjects_list = []
    for s in subjects:
        sc = TYPE_MAP.get(s.subject_code[2], '') if len(s.subject_code) == 6 else ''
        subjects_list.append({'id': s.pk, 'name': s.name, 'category': s.category, 'classification': sc})

    error_message = None

    if request.method == 'POST':
        year = request.POST.get('year', '').strip()
        semester = request.POST.get('semester', '').strip()
        grade_count = int(request.POST.get('grade_count', 0))

        if not year or not semester:
            error_message = '학년과 학기를 선택해 주세요.'
        else:
            created_count = 0
            for i in range(grade_count):
                subject_id = request.POST.get(f'grades[{i}][subject]', '').strip()
                score_str = request.POST.get(f'grades[{i}][score]', '').strip()
                credits_str = request.POST.get(f'grades[{i}][credits]', '').strip()
                enrolled_str = request.POST.get(f'grades[{i}][enrolled_count]', '').strip()

                if not subject_id or not score_str or not credits_str or not enrolled_str:
                    continue
                try:
                    def _int(key):
                        v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                        return int(v) if v else None

                    def _dec(key):
                        v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                        return Decimal(v) if v else None

                    subject = Subject.objects.get(pk=subject_id)
                    Grade.objects.create(
                        student=student,
                        grade_type='internal',
                        subject=subject,
                        year=int(year),
                        semester=int(semester),
                        credits=_int('credits'),
                        score=Decimal(score_str),
                        subject_average=None,
                        subject_stddev=None,
                        grade_rank=_int('grade_rank'),
                        subject_classification='general',
                        is_elective=False,
                        enrolled_count=_int('enrolled_count'),
                        subject_rank=_int('subject_rank'),
                        same_rank_count=_int('same_rank_count'),
                        achievement_level=request.POST.get(f'grades[{i}][achievement_level]', '').strip() or None,
                        percentile=_dec('percentile'),
                    )
                    created_count += 1
                except Exception:
                    continue
            if created_count > 0:
                return redirect('parent_grades', student_pk=student_pk)
            error_message = '과목, 점수, 단위수, 수강자수를 입력해 주세요.'

    context = {
        'student': student,
        'subjects_list': subjects_list,
        'error_message': error_message,
    }
    return render(request, 'students/parent_mock_grade.html', context)


def parent_grade_import(request, student_pk):
    """학부모용 성적 파일 업로드 (세션 인증)"""
    from grades.views import process_csv_file, process_excel_file

    student = _parent_auth(request, student_pk)
    if not student:
        return redirect('parent_lookup')

    error_message = None
    success_message = None

    if request.method == 'POST':
        grade_type = request.POST.get('grade_type', 'internal')
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            error_message = '파일을 선택해 주세요.'
        else:
            file_name = uploaded_file.name.lower()
            try:
                if file_name.endswith('.csv'):
                    result = process_csv_file(uploaded_file, student, grade_type)
                elif file_name.endswith(('.xlsx', '.xls')):
                    result = process_excel_file(uploaded_file, student, grade_type)
                else:
                    error_message = 'CSV 또는 Excel 파일만 업로드 가능합니다.'
                    result = None
                if result:
                    if result.get('created_count', 0) > 0:
                        success_message = f"{result['created_count']}개의 성적이 등록되었습니다."
                    row_errors = result.get('errors', [])
                    if row_errors:
                        error_message = f"{len(row_errors)}개 행에서 오류: {', '.join(row_errors[:3])}"
                    if not result.get('success'):
                        error_message = result.get('message', '처리 중 오류가 발생했습니다.')
            except Exception as e:
                error_message = f'파일 처리 중 오류: {str(e)}'

    is_middle = student.grade in ['K7', 'K8', 'K9']
    context = {
        'student': student,
        'error_message': error_message,
        'success_message': success_message,
        'is_middle': is_middle,
    }
    return render(request, 'students/parent_grade_import.html', context)


def parent_grade_edit(request, student_pk, grade_pk):
    """학부모용 성적 수정 (세션 인증)"""
    from grades.models import Grade
    from decimal import Decimal, InvalidOperation

    student = _parent_auth(request, student_pk)
    if not student:
        return redirect('parent_lookup')

    grade = get_object_or_404(Grade, pk=grade_pk, student=student)
    is_middle = student.grade in ['K7', 'K8', 'K9']
    error_message = None

    if request.method == 'POST':
        try:
            def _dec(key):
                v = request.POST.get(key, '').strip()
                return Decimal(v) if v else None

            def _int(key):
                v = request.POST.get(key, '').strip()
                return int(v) if v else None

            grade.score = _dec('score')
            if grade.score is None:
                raise ValueError('원점수는 필수입니다.')
            grade.subject_average = _dec('subject_average')
            grade.subject_stddev = _dec('subject_stddev')
            grade.enrolled_count = _int('enrolled_count')
            achievement = request.POST.get('achievement_level', '').strip().upper()
            grade.achievement_level = achievement if achievement else None
            grade.save()
            return redirect('parent_grades', student_pk=student_pk)
        except (ValueError, InvalidOperation) as e:
            error_message = str(e)

    context = {
        'student': student,
        'grade': grade,
        'is_middle': is_middle,
        'error_message': error_message,
    }
    return render(request, 'students/parent_grade_edit.html', context)


@login_required
def tuition_certificate_pdf(request, pk):
    """학원교육비(수강료) 납입증명서 PDF 생성"""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    from classes.models import TuitionPayment
    from django.conf import settings
    import os

    student = get_object_or_404(Student, pk=pk)
    year = int(request.GET.get('year', datetime.date.today().year))
    today_date = datetime.date.today()

    # 해당 연도의 월별 납부 금액 집계
    payments = TuitionPayment.objects.filter(
        enrollment__student=student,
        year=year,
    ).values('month').annotate(total=models.Sum('amount'))

    monthly_amounts = {p['month']: p['total'] for p in payments}
    annual_total = sum(monthly_amounts.values())

    # 한글 폰트 등록 — 프로젝트 번들 폰트 우선, Linux/Windows 시스템 폰트 순
    font_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NanumGothic.ttf'),
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        'C:/Windows/Fonts/malgun.ttf',
    ]
    font_name = 'Helvetica'  # fallback
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('KoreanFont', fp))
                font_name = 'KoreanFont'
                break
            except Exception:
                continue

    buffer = io.BytesIO()
    w, h = A4  # 210mm x 297mm

    c = canvas.Canvas(buffer, pagesize=A4)

    def font(size, bold=False):
        c.setFont(font_name, size)

    # 여백
    ml = 20 * mm
    mr = w - 20 * mm
    mt = h - 15 * mm

    # 상단 서식 번호
    font(7)
    c.drawString(ml, mt, '[별지 제44호 서식(2)](\'00.4.3신설)')

    # 외곽선
    top = mt - 6 * mm
    bottom = 20 * mm
    c.setLineWidth(1.5)
    c.rect(ml, bottom, mr - ml, top - bottom)

    # 제목
    font(16)
    title = '학원교육비(수강료) 납입증명서'
    c.drawCentredString(w / 2, top - 12 * mm, title)

    # 제목 아래 구분선
    c.setLineWidth(0.5)
    y = top - 16 * mm
    c.line(ml, y, mr, y)

    # 섹션 1: 신청인
    font(9)
    c.drawString(ml + 2 * mm, y - 6 * mm, '1. 신청인')
    y1 = y - 10 * mm
    c.line(ml, y1, mr, y1)

    # 신청인 표 (① 성명 / ② 주민등록번호)
    row_h = 8 * mm
    mid = ml + (mr - ml) / 2

    c.rect(ml, y1 - row_h, mr - ml, row_h)
    c.line(mid, y1 - row_h, mid, y1)
    # 수직 구분: 성명 라벨/값
    name_label_x = ml + 25 * mm
    c.line(name_label_x, y1 - row_h, name_label_x, y1)
    ssn_label_x = mid + 30 * mm
    c.line(ssn_label_x, y1 - row_h, ssn_label_x, y1)
    font(8)
    c.drawString(ml + 2 * mm, y1 - 5.5 * mm, '① 성명')
    c.drawString(mid + 2 * mm, y1 - 5.5 * mm, '② 주민등록번호')

    y2 = y1 - row_h
    # ③ 주소 행
    c.rect(ml, y2 - row_h, mr - ml, row_h)
    addr_label_x = ml + 25 * mm
    c.line(addr_label_x, y2 - row_h, addr_label_x, y2)
    c.drawString(ml + 2 * mm, y2 - 5.5 * mm, '③ 주소')

    y3 = y2 - row_h
    # 대상 학원생 (2행)
    subject_rows = 2
    subject_h = row_h * subject_rows
    c.rect(ml, y3 - subject_h, mr - ml, subject_h)
    # 대상 학원생 라벨
    subject_label_x = ml + 25 * mm
    c.line(subject_label_x, y3 - subject_h, subject_label_x, y3)
    c.drawString(ml + 2 * mm, y3 - subject_h / 2 - 2 * mm, '대상 학원생')

    # 내부 행 구분선
    c.line(subject_label_x, y3 - row_h, mr, y3 - row_h)
    # ④ 성명 / ⑤ 주민등록번호
    mid2 = subject_label_x + (mr - subject_label_x) / 2
    c.line(mid2, y3 - row_h, mid2, y3)
    name4_x = subject_label_x + 25 * mm
    c.line(name4_x, y3 - row_h, name4_x, y3)
    ssn5_x = mid2 + 30 * mm
    c.line(ssn5_x, y3 - row_h, ssn5_x, y3)

    c.drawString(subject_label_x + 2 * mm, y3 - 5.5 * mm, '④ 성명')
    # 학생 이름
    font(9)
    c.drawString(name4_x + 3 * mm, y3 - 5.5 * mm, student.name)
    font(8)
    c.drawString(mid2 + 2 * mm, y3 - 5.5 * mm, '⑤ 주민등록번호')

    # ⑥ 주소 / ⑦ 소득자와의 관계
    name6_x = subject_label_x + 25 * mm          # ⑥ 라벨 | 내용 구획선
    mid3 = subject_label_x + (mr - subject_label_x) * 0.6  # ⑥영역 | ⑦영역 구획선
    name7_x = mid3 + 35 * mm                      # ⑦ 라벨 | 내용 구획선
    c.line(name6_x, y3 - subject_h, name6_x, y3 - row_h)
    c.line(mid3,    y3 - subject_h, mid3,    y3 - row_h)
    c.line(name7_x, y3 - subject_h, name7_x, y3 - row_h)
    c.drawString(subject_label_x + 2 * mm, y3 - row_h - 5.5 * mm, '⑥ 주소')
    c.drawString(mid3 + 2 * mm, y3 - row_h - 5.5 * mm, '⑦ 소득자와의 관계')

    # 섹션 2: 수강 학원
    y4 = y3 - subject_h
    c.setLineWidth(0.5)
    c.line(ml, y4, mr, y4)
    font(9)
    c.drawString(ml + 2 * mm, y4 - 6 * mm, '2. 수강 학원')
    y5 = y4 - 10 * mm
    c.line(ml, y5, mr, y5)

    # ⑧ 학원명 행
    c.rect(ml, y5 - row_h, mr - ml, row_h)
    lbl8_x = ml + 25 * mm
    c.line(lbl8_x, y5 - row_h, lbl8_x, y5)
    mid8 = lbl8_x + (mr - lbl8_x) * 0.5
    c.line(mid8, y5 - row_h, mid8, y5)
    ssn9_x = mid8 + 30 * mm
    c.line(ssn9_x, y5 - row_h, ssn9_x, y5)
    font(8)
    c.drawString(ml + 2 * mm, y5 - 5.5 * mm, '⑧ 학원명')
    font(9)
    c.drawString(lbl8_x + 3 * mm, y5 - 5.5 * mm, '엠클래스수학과학전문학원')
    font(8)
    c.drawString(mid8 + 2 * mm, y5 - 5.5 * mm, '⑨ 사업자등록번호')
    font(9)
    c.drawString(ssn9_x + 3 * mm, y5 - 5.5 * mm, '134-92-52806')

    y6 = y5 - row_h
    # ⑩ 소재지 행
    c.rect(ml, y6 - row_h, mr - ml, row_h)
    lbl10_x = ml + 25 * mm
    c.line(lbl10_x, y6 - row_h, lbl10_x, y6)
    tel_x = mr - 50 * mm
    c.line(tel_x, y6 - row_h, tel_x, y6)
    tel_lbl_x = tel_x + 18 * mm
    c.line(tel_lbl_x, y6 - row_h, tel_lbl_x, y6)
    font(8)
    c.drawString(ml + 2 * mm, y6 - 5.5 * mm, '⑩ 소재지')
    font(8.5)
    c.drawString(lbl10_x + 3 * mm, y6 - 5.5 * mm, '경기도 안산시 단원구 광덕대로 130 폴리타운 516호')
    font(8)
    c.drawString(tel_x + 2 * mm, y6 - 5.5 * mm, '⑪ 전화번호')
    font(9)
    c.drawString(tel_lbl_x + 2 * mm, y6 - 5.5 * mm, '031-439-1222')

    y7 = y6 - row_h
    # ⑫ 1일 수업시간 / ⑬ 1주간 수업일수 행
    c.rect(ml, y7 - row_h, mr - ml, row_h)
    lbl12_x = ml + 35 * mm
    c.line(lbl12_x, y7 - row_h, lbl12_x, y7)
    val12_x = lbl12_x + 25 * mm
    c.line(val12_x, y7 - row_h, val12_x, y7)
    lbl13_x = val12_x + 35 * mm
    c.line(lbl13_x, y7 - row_h, lbl13_x, y7)
    font(8)
    c.drawString(ml + 2 * mm, y7 - 5.5 * mm, '⑫ 1일 수업시간')
    font(9)
    c.drawString(lbl12_x + 3 * mm, y7 - 5.5 * mm, '2')
    font(8)
    c.drawString(lbl12_x + 8 * mm, y7 - 5.5 * mm, '시간')
    c.drawString(val12_x + 2 * mm, y7 - 5.5 * mm, '⑬ 1주간 수업일수')
    font(9)
    c.drawString(lbl13_x + 3 * mm, y7 - 5.5 * mm, '3')
    font(8)
    c.drawString(lbl13_x + 8 * mm, y7 - 5.5 * mm, '회')

    # 섹션 3: 수강료 납입 금액
    y8 = y7 - row_h
    c.line(ml, y8, mr, y8)
    font(9)
    c.drawString(ml + 2 * mm, y8 - 6 * mm, '3. 수강료 납입 금액')
    y9 = y8 - 10 * mm
    c.line(ml, y9, mr, y9)

    # 월별 납입 금액 표
    table_mid = (ml + mr) / 2
    months_per_col = 6
    col_w = (mr - ml) / 2
    month_lbl_w = 20 * mm
    month_row_h = 8 * mm

    # 헤더 행
    c.rect(ml, y9 - month_row_h, mr - ml, month_row_h)
    c.line(table_mid, y9 - month_row_h, table_mid, y9)
    c.line(ml + month_lbl_w, y9 - month_row_h, ml + month_lbl_w, y9)
    c.line(table_mid + month_lbl_w, y9 - month_row_h, table_mid + month_lbl_w, y9)
    font(8)
    c.drawCentredString(ml + month_lbl_w / 2, y9 - 5.5 * mm, '⑭ 월별')
    c.drawCentredString(ml + month_lbl_w + (col_w - month_lbl_w) / 2, y9 - 5.5 * mm, '⑮ 납입금액')
    c.drawCentredString(table_mid + month_lbl_w / 2, y9 - 5.5 * mm, '⑭ 월별')
    c.drawCentredString(table_mid + month_lbl_w + (col_w - month_lbl_w) / 2, y9 - 5.5 * mm, '⑮ 납입금액')

    y10 = y9 - month_row_h
    for i in range(months_per_col):
        month_l = i + 1
        month_r = i + 7
        row_y = y10 - i * month_row_h

        c.rect(ml, row_y - month_row_h, col_w, month_row_h)
        c.rect(table_mid, row_y - month_row_h, col_w, month_row_h)
        c.line(ml + month_lbl_w, row_y - month_row_h, ml + month_lbl_w, row_y)
        c.line(table_mid + month_lbl_w, row_y - month_row_h, table_mid + month_lbl_w, row_y)

        font(9)
        c.drawCentredString(ml + month_lbl_w / 2, row_y - 5.5 * mm, f'{month_l}월')
        c.drawCentredString(table_mid + month_lbl_w / 2, row_y - 5.5 * mm, f'{month_r}월')

        amt_l = monthly_amounts.get(month_l)
        amt_r = monthly_amounts.get(month_r)
        if amt_l:
            c.drawRightString(mr / 2 - 3 * mm, row_y - 5.5 * mm, f'{amt_l:,}원')
        if amt_r:
            c.drawRightString(mr - 3 * mm, row_y - 5.5 * mm, f'{amt_r:,}원')

    y11 = y10 - months_per_col * month_row_h
    # 연간합계액 행
    c.rect(ml, y11 - month_row_h, mr - ml, month_row_h)
    c.line(ml + month_lbl_w, y11 - month_row_h, ml + month_lbl_w, y11)
    sum_mid = (ml + month_lbl_w + mr) / 2
    c.line(sum_mid, y11 - month_row_h, sum_mid, y11)
    yodo_x = sum_mid + (mr - sum_mid) / 3
    c.line(yodo_x, y11 - month_row_h, yodo_x, y11)
    font(8)
    c.drawCentredString(ml + month_lbl_w / 2, y11 - 5.5 * mm, '연간합계액')
    if annual_total:
        font(9)
        c.drawRightString(sum_mid - 3 * mm, y11 - 5.5 * mm, f'{annual_total:,}원')
    font(8)
    c.drawCentredString(yodo_x - (yodo_x - sum_mid) / 2, y11 - 5.5 * mm, '용도')

    # 서명 이미지 경로 (media/signature.png 또는 .jpg 등)
    sig_path = None
    for ext in ('png', 'jpg', 'jpeg'):
        candidate = os.path.join(settings.MEDIA_ROOT, f'signature.{ext}')
        if os.path.exists(candidate):
            sig_path = candidate
            break

    # 신청인 서명 문구
    y12 = y11 - month_row_h - 4 * mm
    font(8.5)
    from reportlab.lib.utils import simpleSplit
    req_text = ('소득세법 제52조 및 소득세법 시행령 제113조 제1항의 규정에 의하여 교육비 공제를 받고자 하니 '
                '위와 같이 학원교육비(수강료)를 납입하였음을 증명하여 주시기 바랍니다.')
    req_lines = simpleSplit(req_text, font_name, 8.5, mr - ml - 4 * mm)
    for i, line in enumerate(req_lines):
        c.drawString(ml + 2 * mm, y12 - i * 5 * mm, line)
    sig_top_offset = len(req_lines) * 5 * mm
    date_str = f'{today_date.year}년  {today_date.month}월  {today_date.day}일'
    c.drawString(w / 2 - 15 * mm, y12 - sig_top_offset - 4 * mm, date_str)
    c.drawString(ml + 2 * mm, y12 - sig_top_offset - 9 * mm,
                 '              신청인                                   (서명 또는 인)')

    c.line(ml, y12 - sig_top_offset - 13 * mm, mr, y12 - sig_top_offset - 13 * mm)

    y13 = y12 - sig_top_offset - 16 * mm
    c.drawString(ml + 2 * mm, y13, '위와 같이 학원교육비(수강료)를 납입하였음을 증명합니다.')
    c.drawString(w / 2 - 15 * mm, y13 - 5 * mm, date_str)
    c.drawString(ml + 2 * mm, y13 - 10 * mm, '              학원장')

    # 서명 이미지 삽입 (있을 경우)
    if sig_path:
        sig_w = 25 * mm
        sig_h = 12 * mm
        sig_x = mr - sig_w - 5 * mm
        sig_y = y13 - 13 * mm
        font(8.5)
        academy_text = '엠클래스수학과학전문학원'
        c.drawRightString(sig_x - 3 * mm, sig_y + sig_h / 2 - 2 * mm, academy_text)
        c.drawImage(sig_path, sig_x, sig_y,
                    width=sig_w, height=sig_h, preserveAspectRatio=True, mask='auto')
    else:
        c.drawString(mr - 40 * mm, y13 - 10 * mm, '직인 (서명)')

    c.line(ml, y13 - 14 * mm, mr, y13 - 14 * mm)

    # 주석
    y14 = y13 - 17 * mm
    font(7.5)
    notice = (
        '※ 소득세법 제52조에 의한 특별공제를 받을 수 있는 학원의 수강료는 초등학교 취학 전 아동이 학원의 '
        '설립·운영에 관한 법률에 의한 학원에서 1일 3시간 이상, 1주 5일 이상 교육을 실시하는 교육과정의 '
        '교습을 받고 지출한 수강료만 해당합니다.'
    )
    from reportlab.lib.utils import simpleSplit
    lines = simpleSplit(notice, font_name, 7.5, (mr - ml - 4 * mm))
    for i, line in enumerate(lines):
        c.drawString(ml + 2 * mm, y14 - i * 4.5 * mm, line)

    c.save()
    buffer.seek(0)

    filename = f'수강료납입증명서_{student.name}_{year}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    return response