import datetime
import io

from django.views.generic import TemplateView
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum, F
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone


class AdminLoginView(LoginView):
    """관리자 전용 로그인 뷰 - 교사 계정 차단"""
    template_name = 'index.html'

    def form_valid(self, form):
        user = form.get_user()
        # 교사 계정인 경우 로그인 차단
        if hasattr(user, 'teacher_profile'):
            messages.error(self.request, '선생님 계정은 "선생님 포털 로그인"을 이용해 주세요.')
            return self.form_invalid(form)
        return super().form_valid(form)


class IndexView(TemplateView):
    template_name = 'index.html'

    # 공개 홈페이지로 서빙할 호스트 목록
    PUBLIC_HOSTS = ('mclass.co.kr', 'www.mclass.co.kr')

    def get(self, request, *args, **kwargs):
        host = request.get_host().split(':')[0].lower()
        # mclass.co.kr 는 로그인 여부·권한 무관하게 항상 공개 홈페이지
        if host in self.PUBLIC_HOSTS:
            from homepage.views import homepage_index
            return homepage_index(request)

        if request.user.is_authenticated:
            # manager.mclass.co.kr 의 로그인된 관리자 → 대시보드
            self.template_name = 'dashboard.html'
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_authenticated:
            from students.models import Student
            from teachers.models import Teacher, TeacherStudentAssignment, TeacherUnavailability, Attendance, Salary
            from bookstore.models import StudentBookProgress, Book, BookStockLog
            from maintenance.models import Maintenance

            today = timezone.now().date()
            this_month = today.replace(day=1)

            # 학년별 학생 수 집계
            grade_counts = Student.objects.filter(
                is_active=True
            ).values('grade').annotate(count=Count('id'))
            grade_dict = {item['grade']: item['count'] for item in grade_counts if item['grade']}
            grade_order = ['K5', 'K6', 'K7', 'K8', 'K9', 'K10', 'K11', 'K12']
            grade_stats = [(grade, grade_dict.get(grade, 0)) for grade in grade_order if grade in grade_dict]
            total_students = Student.objects.filter(is_active=True).count()

            # 재직 중인 교사 수
            active_teachers = Teacher.objects.filter(
                Q(resignation_date__isnull=True)
            ).count()

            # 출근 불가 일정 승인 대기 건수
            pending_unavailability_count = TeacherUnavailability.objects.filter(
                status='pending'
            ).count()

            # 오늘 학습 진도 통계
            today_assignments = TeacherStudentAssignment.objects.filter(date=today)
            today_assigned = today_assignments.filter(assignment_type='normal').count()
            today_absent = today_assignments.filter(assignment_type='absent').count()
            today_progress_records = StudentBookProgress.objects.filter(study_date=today).count()

            # 이번 달 입고 금액 / 미정산 금액
            # created_at__date__gte 는 이번 달 이후 전체를 포함하므로, year+month 로 이번 달만 정확히 필터
            month_logs = BookStockLog.objects.filter(
                created_at__year=today.year,
                created_at__month=today.month,
            )
            inbound = month_logs.filter(quantity__gt=0).aggregate(s=Sum('total_payment'))['s'] or 0
            returned = month_logs.filter(quantity__lt=0).aggregate(s=Sum('total_payment'))['s'] or 0
            month_inbound_payment = inbound - returned
            month_unpaid_payment = month_logs.filter(is_paid=False, quantity__gt=0).aggregate(
                s=Sum('total_payment'))['s'] or 0

            # 재고 부족 교재 (3권 이하)
            low_stock_count = Book.objects.filter(stock__lte=3).count()

            # 미납 학생 수 (unpaid_amount > 0)
            unpaid_student_count = Student.objects.filter(is_active=True, unpaid_amount__gt=0).count()

            # 오늘 출근 교사 수
            today_present = Attendance.objects.filter(date=today, is_present=True).count()

            # 이번 달 수입 - 수강료 / 교재 판매
            from classes.models import MonthlyEnrollment, TuitionPayment
            from bookstore.models import BookSale
            from django.db.models import F

            # 이번 달 MonthlyEnrollment (취소 제외)
            monthly_this_month = list(
                MonthlyEnrollment.objects.filter(
                    year=today.year,
                    month=today.month,
                ).exclude(status='cancelled').select_related('lesson')
            )

            # 1) 이번 달 청구 금액 (MonthlyEnrollment 합산)
            month_tuition_billing = sum(me.adjusted_tuition for me in monthly_this_month)

            # 1-1) 전월 중간 등록 학생의 전월 미납 수강료 (billing export의 prev_month_fee와 동일)
            # billing export는 전월 중간 등록(enrollment_date__day > 1) 학생의 전월 미납분을 이번 달 청구에 포함
            from classes.models import Enrollment as _Enrollment
            if today.month == 1:
                prev_year_val, prev_month_val = today.year - 1, 12
            else:
                prev_year_val, prev_month_val = today.year, today.month - 1

            # 이번 달 청구 대상 학생 PK
            # billing export와 동일하게: 이번 달 수강 학생 + 미납 교재 있는 학생 전체
            this_month_student_pks = set(me.student_id for me in monthly_this_month)
            unpaid_book_student_pks = set(
                BookSale.objects.filter(is_paid=False).values_list('student_id', flat=True)
            )
            all_billing_student_pks = this_month_student_pks | unpaid_book_student_pks

            # 전월 중간(2일 이후) 등록한 (student_id, lesson_id)
            mid_month_keys = set(
                _Enrollment.objects.filter(
                    student_id__in=all_billing_student_pks,
                    enrollment_date__year=prev_year_val,
                    enrollment_date__month=prev_month_val,
                    enrollment_date__day__gt=1,
                ).values_list('student_id', 'lesson_id')
            )

            prev_month_fee_total = 0
            if mid_month_keys:
                # 전월 납부 완료 집합
                paid_prev_quads = set(
                    TuitionPayment.objects.filter(
                        year=prev_year_val,
                        month=prev_month_val,
                        enrollment__student_id__in=all_billing_student_pks,
                    ).values_list('enrollment__student_id', 'enrollment__lesson_id', 'year', 'month')
                )
                # 전월 중간 등록 학생의 전월 미납 MonthlyEnrollment
                prev_mes = MonthlyEnrollment.objects.filter(
                    student_id__in=all_billing_student_pks,
                    year=prev_year_val,
                    month=prev_month_val,
                ).exclude(status='cancelled').select_related('lesson')
                prev_month_fee_total = sum(
                    me.adjusted_tuition for me in prev_mes
                    if (me.student_id, me.lesson_id) in mid_month_keys
                    and (me.student_id, me.lesson_id, me.year, me.month) not in paid_prev_quads
                )
                month_tuition_billing += prev_month_fee_total

            # 2) 이번 달 수납 금액 (TuitionPayment 기준)
            month_tuition_collected = TuitionPayment.objects.filter(
                year=today.year,
                month=today.month,
            ).aggregate(s=Sum('amount'))['s'] or 0

            # 2-1) 이번 달 환불 금액 (WithdrawalRefund 기준)
            from classes.models import WithdrawalRefund
            month_tuition_refund = WithdrawalRefund.objects.filter(
                year=today.year,
                month=today.month,
            ).aggregate(s=Sum('refund_amount'))['s'] or 0

            # 3) 이번 달 미납 금액 (청구됐으나 TuitionPayment 없는 MonthlyEnrollment 합산)
            # TuitionPayment → Enrollment → (student_id, lesson_id) 로 cross-reference
            paid_pairs = set(
                TuitionPayment.objects.filter(
                    year=today.year,
                    month=today.month,
                ).values_list('enrollment__student_id', 'enrollment__lesson_id')
            )
            month_tuition_unpaid = sum(
                me.adjusted_tuition for me in monthly_this_month
                if (me.student_id, me.lesson_id) not in paid_pairs
            ) + prev_month_fee_total

            # 교재비 청구 = 누적 판매 총액 (is_paid 무관, 납부해도 변하지 않음)
            _book_qs = BookSale.objects.aggregate(
                billing=Sum(F('price') * F('quantity')),
                collected=Sum(F('price') * F('quantity'), filter=Q(is_paid=True)),
            )
            month_book_billing   = _book_qs['billing']   or 0
            month_book_collected = _book_qs['collected']  or 0
            month_book_unpaid    = month_book_billing - month_book_collected

            # 합계 (청구 기준) = 수강료 청구 + 교재비 미납 누계
            total_billing = month_tuition_billing + month_book_unpaid

            # 이번 달 정산 - 지출
            month_maint = Maintenance.objects.filter(
                date__year=today.year,
                date__month=today.month,
            )
            month_rent = month_maint.aggregate(s=Sum('rent'))['s'] or 0
            month_charge = month_maint.aggregate(s=Sum('charge'))['s'] or 0

            # 교사 급여: SalaryCalculationView 실시간 계산 (추가급여 미입력분은 0으로 처리)
            from teachers.views import SalaryCalculationView
            salary_view = SalaryCalculationView()
            active_teachers_month = salary_view.get_active_teachers_for_month(today.year, today.month)
            month_salary = 0
            for teacher in active_teachers_month:
                work_hours, _ = salary_view.calculate_work_hours(teacher, today.year, today.month)
                base_amount = int(work_hours * teacher.base_salary)
                try:
                    existing = Salary.objects.get(teacher=teacher, year=today.year, month=today.month)
                    additional = existing.additional_amount
                except Salary.DoesNotExist:
                    additional = 0
                month_salary += base_amount + additional

            total_expense = month_rent + month_charge + month_salary + month_inbound_payment

            # 미납 현황 (누적) - BookSale 실시간 계산 기준
            total_unpaid_book = BookSale.objects.filter(is_paid=False).aggregate(
                s=Sum(F('price') * F('quantity')))['s'] or 0
            total_unpaid = total_unpaid_book

            # 다음 달 수강 현황 (MonthlyEnrollment 기준)
            import calendar as _cal
            if today.month == 12:
                next_year_val, next_month_val = today.year + 1, 1
            else:
                next_year_val, next_month_val = today.year, today.month + 1

            next_month_qs = MonthlyEnrollment.objects.filter(
                year=next_year_val,
                month=next_month_val,
            ).select_related('lesson')

            confirmed_next = next_month_qs.filter(status='confirmed')
            next_month_confirmed = confirmed_next.count()
            next_month_confirmed_tuition = sum(me.adjusted_tuition for me in confirmed_next)
            next_month_pending = next_month_qs.filter(status='pending').count()

            context.update({
                'grade_stats': grade_stats,
                'total_students': total_students,
                'active_teachers': active_teachers,
                'pending_unavailability_count': pending_unavailability_count,
                'today': today,
                'today_assigned': today_assigned,
                'today_absent': today_absent,
                'today_progress_records': today_progress_records,
                'month_inbound_payment': month_inbound_payment,
                'month_unpaid_payment': month_unpaid_payment,
                'low_stock_count': low_stock_count,
                'unpaid_student_count': unpaid_student_count,
                'today_present': today_present,
                'month_rent': month_rent,
                'month_charge': month_charge,
                'month_salary': month_salary,
                'total_expense': total_expense,
                'month_tuition_billing': month_tuition_billing,
                'month_tuition_collected': month_tuition_collected,
                'month_tuition_refund': month_tuition_refund,
                'month_tuition_unpaid': month_tuition_unpaid,
                'month_book_billing': month_book_billing,
                'month_book_collected': month_book_collected,
                'month_book_unpaid': month_book_unpaid,
                'total_billing': total_billing,
                'total_unpaid_book': total_unpaid_book,
                'total_unpaid': total_unpaid,
                'next_year_val': next_year_val,
                'next_month_val': next_month_val,
                'next_month_confirmed': next_month_confirmed,
                'next_month_confirmed_tuition': next_month_confirmed_tuition,
                'next_month_pending': next_month_pending,
            })

        return context


@login_required
def billing_export(request):
    """다음 달 수강료 자동 청구용 xlsx 내보내기"""
    if not request.user.is_staff:
        messages.error(request, '관리자만 사용 가능합니다.')
        from django.shortcuts import redirect
        return redirect('index')

    from students.models import Student
    from classes.models import MonthlyEnrollment
    from bookstore.models import BookSale

    today = datetime.date.today()

    # GET/POST 파라미터로 연도·월 선택 (기본값: 이번 달)
    param_year = request.GET.get('year') or request.POST.get('year')
    param_month = request.GET.get('month') or request.POST.get('month')
    if param_year and param_month:
        target_year, target_month = int(param_year), int(param_month)
    else:
        target_year, target_month = today.year, today.month

    # 선택 가능 월 목록: 이번 달 ~ 다음 달
    month_options = []
    base = today.replace(day=1)
    for delta in range(0, 2):
        d = (base + datetime.timedelta(days=32 * delta)).replace(day=1)
        month_options.append({'year': d.year, 'month': d.month})

    # 선택된 달 MonthlyEnrollment (취소 제외)
    next_enrollments = (
        MonthlyEnrollment.objects
        .filter(year=target_year, month=target_month)
        .exclude(status='cancelled')
        .select_related('student', 'lesson')
    )

    # 미납 교재 판매
    unpaid_sales = (
        BookSale.objects
        .filter(is_paid=False)
        .select_related('student', 'book')
    )

    # 학생별로 집계
    from collections import defaultdict
    student_map = {}  # pk -> dict

    def get_or_create(student):
        if student.pk not in student_map:
            student_map[student.pk] = {
                'student': student,
                'tuition_items': [],   # (lesson_name, amount)
                'book_items': [],      # (book_title, amount)
            }
        return student_map[student.pk]

    for enr in next_enrollments:
        entry = get_or_create(enr.student)
        entry['tuition_items'].append((enr.lesson.name, enr.adjusted_tuition))

    # 해당 월에 운영 중인 특별 수업 (미납) 포함
    import calendar as _cal
    last_day = _cal.monthrange(target_year, target_month)[1]
    first_of_month = datetime.date(target_year, target_month, 1)
    last_of_month  = datetime.date(target_year, target_month, last_day)

    from classes.models import Enrollment as ClassEnrollment, TuitionPayment as _TP
    special_enrollments = ClassEnrollment.objects.filter(
        is_active=True,
        lesson__is_special=True,
        lesson__start_date__lte=last_of_month,
        lesson__end_date__gte=first_of_month,
    ).select_related('student', 'lesson')

    paid_special_ids = set(
        _TP.objects.filter(
            enrollment__lesson__is_special=True,
        ).values_list('enrollment_id', flat=True)
    )

    for enr in special_enrollments:
        if enr.pk not in paid_special_ids:
            entry = get_or_create(enr.student)
            entry.setdefault('special_items', []).append(
                (enr.lesson.name, enr.adjusted_tuition)
            )

    for sale in unpaid_sales:
        entry = get_or_create(sale.student)
        entry['book_items'].append((sale.book.title, sale.price * sale.quantity))

    # 합계 계산 후 이름 순 정렬
    for entry in student_map.values():
        entry.setdefault('special_items', [])
        entry['tuition_total'] = sum(amt for _, amt in entry['tuition_items'])
        entry['special_total'] = sum(amt for _, amt in entry['special_items'])
        entry['book_total'] = sum(amt for _, amt in entry['book_items'])
        entry['total'] = entry['tuition_total'] + entry['special_total'] + entry['book_total']

    rows = sorted(student_map.values(), key=lambda x: x['student'].name)

    # 전월 미납 수강료 계산 (target_year/month 이전)
    from classes.models import TuitionPayment
    from django.db.models import Q as _Q

    student_pks = list(student_map.keys())

    # 대상 학생들의 납부 완료 (lesson_id, year, month) 집합
    paid_quads = set(
        TuitionPayment.objects
        .filter(enrollment__student_id__in=student_pks)
        .values_list('enrollment__student_id', 'enrollment__lesson_id', 'year', 'month')
    )

    # target 월 이전 미납 MonthlyEnrollment
    past_mes = (
        MonthlyEnrollment.objects
        .filter(student_id__in=student_pks)
        .filter(_Q(year__lt=target_year) | _Q(year=target_year, month__lt=target_month))
        .exclude(status='cancelled')
        .select_related('lesson')
        .order_by('year', 'month', 'lesson__name')
    )

    # 전월 계산 (prev_month는 template 컨텍스트용으로만 유지)
    if target_month == 1:
        prev_month, prev_year = 12, target_year - 1
    else:
        prev_month, prev_year = target_month - 1, target_year

    # 모든 과거 미납: 청구금액·내용에 반영
    all_unpaid_dict = {}  # student_pk -> [{'label': str, 'name': str, 'amount': int}, ...]

    for me in past_mes:
        if (me.student_id, me.lesson_id, me.year, me.month) not in paid_quads:
            label = f'{me.month}월' if me.year == target_year else f'{me.year}년 {me.month}월'
            all_unpaid_dict.setdefault(me.student_id, []).append(
                {'label': label, 'name': me.lesson.name, 'amount': me.adjusted_tuition}
            )

    HOMEPAGE_MSG = '자세한 내역은 학원 홈페이지(https://mclass.co.kr)에서 확인할 수 있습니다.'

    for entry in rows:
        student = entry['student']
        unpaid_items = all_unpaid_dict.get(student.pk, [])
        unpaid_total = sum(item['amount'] for item in unpaid_items)

        entry['unpaid_items'] = unpaid_items
        entry['unpaid_total'] = unpaid_total

        memo_parts = [f'고유 번호 : {student.student_id}'] if student.student_id else []
        memo_parts.append(HOMEPAGE_MSG)
        entry['auto_memo'] = ' / '.join(memo_parts)

    # POST → 파일 생성
    if request.method == 'POST':
        suffix = request.POST.get('file_suffix', '1st').strip() or '1st'
        export_format = request.POST.get('export_format', 'xlsx')

        # 체크된 학생만 필터링 (체크박스 미선택 시 전체)
        selected_pks = request.POST.getlist('student_pks')
        if selected_pks:
            selected_pks_set = set(int(pk) for pk in selected_pks)
            export_rows = [r for r in rows if r['student'].pk in selected_pks_set]
        else:
            export_rows = rows

        # 공통 데이터 행 생성
        headers = ['이름', '부모 전화번호', '청구금액', '내용', '메모']
        data_rows = []
        for entry in export_rows:
            student = entry['student']
            parts = []
            for item in entry['unpaid_items']:
                parts.append(f'{item["label"]} 미납 {item["name"]} {item["amount"]:,}원')
            for name, amt in entry['tuition_items']:
                parts.append(f'{name} {amt:,}원')
            for name, amt in entry['special_items']:
                parts.append(f'{name} {amt:,}원')
            if entry['book_total'] > 0:
                parts.append(f'교재비 {entry["book_total"]:,}원')
            content = ', '.join(parts)
            data_rows.append([
                student.name,
                student.parent_phone or '',
                entry['total'] + entry['unpaid_total'],
                content,
                entry['auto_memo'],
            ])

        if export_format == 'csv':
            import csv as _csv
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            filename = f'payssam_format_{target_year}_{target_month:02d}_{suffix}.csv'
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
            writer = _csv.writer(response)
            writer.writerow(headers)
            writer.writerows(data_rows)
            return response

        # xlsx
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{target_year}년 {target_month}월 청구'

        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        thin = Side(border_style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, values in enumerate(data_rows, 2):
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=(col == 4))

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'payssam_format_{target_year}_{target_month:02d}_{suffix}.xlsx'
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{filename}'
        return response

    # GET → 미리보기 페이지
    context = {
        'rows': rows,
        'next_year': target_year,
        'next_month': target_month,
        'target_year': target_year,
        'target_month': target_month,
        'prev_month': prev_month,
        'month_options': month_options,
    }
    return render(request, 'billing_export.html', context)
