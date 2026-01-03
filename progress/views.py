# progress/views.py

import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse, HttpResponse
from django.db import transaction

from .models import ProblemType, BookProblem, StudentProgress, ProgressEntry
from .forms import (
    ProblemTypeForm, ProblemTypeUploadForm,
    BookProblemForm, ProgressEntryForm, StudentProgressCreateForm
)
from bookstore.models import Book, BookSale
from teachers.models import Teacher
from subjects.models import Subject


# ========================
# 문제 유형 (ProblemType) 관련 뷰
# ========================

class ProblemTypeListView(ListView):
    """문제 유형 목록"""
    model = ProblemType
    template_name = 'progress/problem_type_list.html'
    context_object_name = 'problem_types'
    paginate_by = 50

    def get_queryset(self):
        queryset = ProblemType.objects.all()
        subject_code = self.request.GET.get('subject')
        if subject_code:
            queryset = queryset.filter(code_number__startswith=subject_code)
        return queryset.order_by('code_number')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['subjects'] = Subject.objects.filter(is_active=True).order_by('subject_code')
        context['selected_subject'] = self.request.GET.get('subject', '')
        return context


class ProblemTypeCreateView(View):
    """문제 유형 개별 등록"""
    template_name = 'progress/problem_type_form.html'

    def get(self, request):
        form = ProblemTypeForm()
        subjects = Subject.objects.filter(is_active=True).order_by('subject_code')
        return render(request, self.template_name, {
            'form': form,
            'subjects': subjects,
            'title': '문제 유형 등록'
        })

    def post(self, request):
        form = ProblemTypeForm(request.POST)
        subjects = Subject.objects.filter(is_active=True).order_by('subject_code')
        if form.is_valid():
            form.save()
            messages.success(request, '문제 유형이 등록되었습니다.')
            return redirect('progress:problem_type_list')
        return render(request, self.template_name, {
            'form': form,
            'subjects': subjects,
            'title': '문제 유형 등록'
        })


class ProblemTypeUploadView(View):
    """CSV/Excel 파일로 문제 유형 일괄 업로드"""
    template_name = 'progress/problem_type_upload.html'

    def get(self, request):
        form = ProblemTypeUploadForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = ProblemTypeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']
            ext = file.name.split('.')[-1].lower()

            try:
                if ext == 'csv':
                    problem_types = self._parse_csv(file)
                else:
                    problem_types = self._parse_excel(file)

                created_count = 0
                updated_count = 0
                error_count = 0
                errors = []

                with transaction.atomic():
                    for pt_data in problem_types:
                        code_number = str(pt_data.get('code_number', '')).strip()

                        # 17자리 코드 번호 검증
                        if not code_number or len(code_number) != 17 or not code_number.isdigit():
                            error_count += 1
                            errors.append(f"잘못된 코드 번호: {code_number} (17자리 숫자여야 합니다)")
                            continue

                        # 과목 코드 확인
                        subject_code = code_number[:4]
                        if not Subject.objects.filter(subject_code=subject_code).exists():
                            error_count += 1
                            errors.append(f"존재하지 않는 과목 코드: {subject_code} (코드: {code_number})")
                            continue

                        # 난도 처리 (기본값: 5)
                        difficulty_raw = pt_data.get('difficulty', '5')
                        try:
                            difficulty = int(difficulty_raw) if difficulty_raw else 5
                            if not (1 <= difficulty <= 10):
                                difficulty = 5
                        except (ValueError, TypeError):
                            difficulty = 5

                        obj, created = ProblemType.objects.update_or_create(
                            code_number=code_number,
                            defaults={
                                'title': pt_data.get('title', ''),
                                'memo': pt_data.get('memo', ''),
                                'difficulty': difficulty,
                            }
                        )
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                msg = f"업로드 완료: {created_count}개 생성, {updated_count}개 업데이트"
                if error_count > 0:
                    msg += f", {error_count}개 오류"
                messages.success(request, msg)

                if errors:
                    for err in errors[:5]:  # 최대 5개 에러만 표시
                        messages.warning(request, err)

                return redirect('progress:problem_type_list')

            except Exception as e:
                messages.error(request, f"파일 처리 중 오류가 발생했습니다: {str(e)}")

        return render(request, self.template_name, {'form': form})

    def _parse_csv(self, file):
        """CSV 파일 파싱"""
        decoded_file = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded_file))
        return list(reader)

    def _parse_excel(self, file):
        """Excel 파일 파싱"""
        try:
            import openpyxl
        except ImportError:
            raise Exception("openpyxl 라이브러리가 필요합니다.")

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)

        return rows


class ProblemTypeDeleteView(View):
    """문제 유형 삭제"""

    def post(self, request, pk):
        problem_type = get_object_or_404(ProblemType, pk=pk)
        problem_type.delete()
        messages.success(request, '문제 유형이 삭제되었습니다.')
        return redirect('progress:problem_type_list')


# ========================
# 교재-문제 연결 (BookProblem) 관련 뷰
# ========================

class BookProblemListView(ListView):
    """교재별 문제 목록"""
    model = BookProblem
    template_name = 'progress/book_problem_list.html'
    context_object_name = 'problems'

    def get_queryset(self):
        queryset = BookProblem.objects.select_related('book', 'problem_type')
        book_id = self.request.GET.get('book')
        if book_id:
            queryset = queryset.filter(book_id=book_id)
        return queryset.order_by('book', 'order')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['books'] = Book.objects.all().order_by('title')
        context['selected_book'] = self.request.GET.get('book', '')
        return context


class BookProblemCreateView(CreateView):
    """교재-문제 유형 연결"""
    model = BookProblem
    form_class = BookProblemForm
    template_name = 'progress/book_problem_form.html'
    success_url = reverse_lazy('progress:book_problem_list')

    def form_valid(self, form):
        messages.success(self.request, '교재에 문제가 연결되었습니다.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '교재-문제 연결'
        return context


class BookProblemDeleteView(View):
    """교재-문제 연결 삭제"""

    def post(self, request, pk):
        problem = get_object_or_404(BookProblem, pk=pk)
        book_id = problem.book_id
        problem.delete()
        messages.success(request, '연결이 삭제되었습니다.')
        return redirect(f"{reverse('progress:book_problem_list')}?book={book_id}")


# ========================
# 학생 진도표 관련 뷰
# ========================

class StudentProgressListView(ListView):
    """학생 진도표 목록"""
    model = StudentProgress
    template_name = 'progress/student_progress_list.html'
    context_object_name = 'progress_list'

    def get_queryset(self):
        return StudentProgress.objects.select_related(
            'student', 'book', 'book_sale'
        ).order_by('-created_at')


class StudentProgressCreateView(View):
    """학생 진도표 생성 (별도 버튼으로 수행)"""
    template_name = 'progress/student_progress_create.html'

    def get(self, request):
        form = StudentProgressCreateForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = StudentProgressCreateForm(request.POST)
        if form.is_valid():
            book_sale = form.cleaned_data['book_sale']

            with transaction.atomic():
                progress = StudentProgress.objects.create(
                    student=book_sale.student,
                    book_sale=book_sale,
                    book=book_sale.book
                )

                problems = book_sale.book.problems.all().order_by('order')
                entries = [
                    ProgressEntry(progress=progress, book_problem=problem)
                    for problem in problems
                ]
                ProgressEntry.objects.bulk_create(entries)

            messages.success(
                request,
                f"'{book_sale.student.name}'의 '{book_sale.book.title}' 진도표가 생성되었습니다. ({len(entries)}개 항목)"
            )
            return redirect('progress:student_progress_detail', pk=progress.pk)

        return render(request, self.template_name, {'form': form})


class StudentProgressDetailView(DetailView):
    """학생 진도표 상세 (진도 항목 목록)"""
    model = StudentProgress
    template_name = 'progress/student_progress_detail.html'
    context_object_name = 'progress'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['entries'] = self.object.entries.select_related(
            'book_problem', 'book_problem__problem_type', 'teacher'
        ).order_by('book_problem__order')
        context['teachers'] = Teacher.objects.filter(is_active=True).order_by('name')
        context['understanding_choices'] = ProgressEntry.UNDERSTANDING_CHOICES
        return context


class ProgressEntryUpdateView(View):
    """진도 항목 AJAX 업데이트"""

    def post(self, request, pk):
        entry = get_object_or_404(ProgressEntry, pk=pk)

        is_assigned = request.POST.get('is_assigned')
        if is_assigned is not None:
            entry.is_assigned = is_assigned == 'true'

        study_date = request.POST.get('study_date')
        if study_date is not None:
            entry.study_date = study_date if study_date else None

        teacher_id = request.POST.get('teacher')
        if teacher_id is not None:
            entry.teacher_id = teacher_id if teacher_id else None

        understanding = request.POST.get('understanding')
        if understanding is not None:
            entry.understanding = int(understanding) if understanding else None

        homework = request.POST.get('homework')
        if homework is not None:
            entry.homework = homework

        memo = request.POST.get('memo')
        if memo is not None:
            entry.memo = memo

        entry.save()

        return JsonResponse({
            'success': True,
            'entry_id': entry.pk,
            'completion_rate': entry.progress.get_completion_rate()
        })


def download_problem_type_template(request):
    """문제 유형 일괄 업로드용 샘플 템플릿 다운로드"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="problem_type_template.csv"'

    # UTF-8 BOM 추가 (Excel에서 한글 깨짐 방지)
    response.write('\ufeff')

    writer = csv.writer(response)
    # 헤더 작성
    writer.writerow(['code_number', 'difficulty', 'title', 'memo'])

    # 샘플 데이터 작성
    writer.writerow(['50010101001001', '5', '1단원 유형1 문제1', '샘플 메모'])
    writer.writerow(['50010101001002', '3', '1단원 유형1 문제2', ''])
    writer.writerow(['50010102001001', '7', '1단원 유형2 문제1', '난이도 높음'])

    return response
