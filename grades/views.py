from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import Case, When, Value, IntegerField
from students.models import Student
from subjects.models import Subject
from .models import Grade
from .forms import InternalGradeForm, MockExamGradeForm, InternalGradeBulkFormSet, GradeImportForm
import csv
import io
import json
from decimal import Decimal, InvalidOperation
from collections import defaultdict


@login_required
def internal_grade_create(request, student_pk):
    """내신 성적 입력"""
    student = get_object_or_404(Student, pk=student_pk)

    if request.method == 'POST':
        form = InternalGradeForm(request.POST)
        if form.is_valid():
            grade = form.save(commit=False)
            grade.student = student
            grade.grade_type = 'internal'
            grade.full_clean()  # 모델 유효성 검사
            grade.save()
            messages.success(request, '내신 성적이 등록되었습니다.')
            return redirect('students:student_detail', pk=student_pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = InternalGradeForm()

    context = {
        'form': form,
        'student': student,
        'grade_type': 'internal',
    }
    return render(request, 'grades/grade_form.html', context)


@login_required
def mock_grade_create(request, student_pk):
    """모의고사 성적 입력"""
    student = get_object_or_404(Student, pk=student_pk)

    if request.method == 'POST':
        form = MockExamGradeForm(request.POST)
        if form.is_valid():
            grade = form.save(commit=False)
            grade.student = student
            grade.grade_type = 'mock'
            grade.full_clean()  # 모델 유효성 검사
            grade.save()
            messages.success(request, '모의고사 성적이 등록되었습니다.')
            return redirect('students:student_detail', pk=student_pk)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = MockExamGradeForm()

    context = {
        'form': form,
        'student': student,
        'grade_type': 'mock',
    }
    return render(request, 'grades/grade_form.html', context)


@login_required
def grade_update(request, pk):
    """성적 수정"""
    grade = get_object_or_404(Grade, pk=pk)
    student = grade.student

    # 성적 유형에 따라 적절한 폼 선택
    if grade.grade_type == 'internal':
        FormClass = InternalGradeForm
        grade_type_label = '내신'
    else:
        FormClass = MockExamGradeForm
        grade_type_label = '모의고사'

    if request.method == 'POST':
        form = FormClass(request.POST, instance=grade)
        if form.is_valid():
            try:
                updated_grade = form.save(commit=False)
                updated_grade.full_clean()
                updated_grade.save()
                messages.success(request, f'{grade_type_label} 성적이 수정되었습니다.')
                return redirect('students:student_detail', pk=student.pk)
            except Exception as e:
                messages.error(request, f'성적 수정 중 오류가 발생했습니다: {str(e)}')
        else:
            # 폼 에러 출력
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FormClass(instance=grade)

    context = {
        'form': form,
        'student': student,
        'grade': grade,
        'grade_type': grade.grade_type,
        'is_update': True,
    }
    return render(request, 'grades/grade_form.html', context)


@login_required
def grade_delete(request, pk):
    """성적 삭제"""
    grade = get_object_or_404(Grade, pk=pk)
    student = grade.student

    if request.method == 'POST':
        grade_type_label = '내신' if grade.grade_type == 'internal' else '모의고사'
        grade.delete()
        messages.success(request, f'{grade_type_label} 성적이 삭제되었습니다.')
        return redirect('students:student_detail', pk=student.pk)

    context = {
        'grade': grade,
        'student': student,
    }
    return render(request, 'grades/grade_confirm_delete.html', context)


@login_required
def internal_grade_bulk_create(request, student_pk):
    """한 학기 내신 성적 일괄 입력"""
    student = get_object_or_404(Student, pk=student_pk)
    is_middle = student.grade in ['K7', 'K8', 'K9']
    is_2022 = (not is_middle) and (student.curriculum_year == 2022)

    if request.method == 'POST':
        try:
            year = request.POST.get('year')
            semester = request.POST.get('semester')
            grade_count = int(request.POST.get('grade_count', 0))

            if not year or not semester:
                messages.error(request, '학년과 학기를 선택해주세요.')
                return redirect('grades:internal_grade_bulk_create', student_pk=student_pk)

            created_count = 0
            for i in range(grade_count):
                subject_id = request.POST.get(f'grades[{i}][subject]')
                if not subject_id:
                    continue

                credits = request.POST.get(f'grades[{i}][credits]')
                score = request.POST.get(f'grades[{i}][score]')

                def _get(key):
                    v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                    return v if v else None

                def _int(key):
                    v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                    return int(v) if v else None

                subject = Subject.objects.get(pk=subject_id)

                if is_middle:
                    grade = Grade(
                        student=student,
                        grade_type='internal',
                        subject=subject,
                        year=year,
                        semester=semester,
                        credits=credits,
                        score=score,
                        subject_average=_get('subject_average'),
                        subject_stddev=_get('subject_stddev'),
                        enrolled_count=_int('enrolled_count'),
                        achievement_level=_get('achievement_level'),
                        subject_classification='general',
                        is_elective=False,
                    )
                elif is_2022:
                    classification_raw = request.POST.get(f'grades[{i}][subject_classification]', 'general')
                    subject_classification = classification_raw if classification_raw in ('common', 'general', 'elective', 'fusion') else 'general'
                    grade = Grade(
                        student=student,
                        grade_type='internal',
                        subject=subject,
                        year=year,
                        semester=semester,
                        credits=credits,
                        score=score,
                        subject_average=_get('subject_average'),
                        subject_stddev=None,
                        grade_rank=_int('grade_rank'),
                        subject_classification=subject_classification,
                        is_elective=(subject_classification in ('elective', 'fusion')),
                        enrolled_count=_int('enrolled_count'),
                        subject_rank=_int('subject_rank'),
                        same_rank_count=_int('same_rank_count'),
                        achievement_level=_get('achievement_level'),
                        percentile=_get('percentile'),
                    )
                else:
                    classification_raw = request.POST.get(f'grades[{i}][subject_classification]', 'general')
                    subject_classification = classification_raw if classification_raw in ('common', 'general', 'elective', 'fusion') else 'general'
                    is_achievement = subject_classification in ('elective', 'fusion')
                    grade_rank_raw = request.POST.get(f'grades[{i}][grade_rank]')
                    def _dec(key):
                        v = request.POST.get(f'grades[{i}][{key}]', '').strip()
                        from decimal import Decimal
                        return Decimal(v) if v else None

                    grade = Grade(
                        student=student,
                        grade_type='internal',
                        subject=subject,
                        year=year,
                        semester=semester,
                        credits=credits,
                        score=score,
                        subject_average=request.POST.get(f'grades[{i}][subject_average]'),
                        subject_stddev=request.POST.get(f'grades[{i}][subject_stddev]') if not is_achievement else None,
                        grade_rank=int(grade_rank_raw) if grade_rank_raw and not is_achievement else None,
                        subject_classification=subject_classification,
                        is_elective=is_achievement,
                        enrolled_count=None,
                        achievement_level=_get('achievement_level') if is_achievement else None,
                        distribution_a=_dec('distribution_a') if is_achievement else None,
                        distribution_b=_dec('distribution_b') if is_achievement else None,
                        distribution_c=_dec('distribution_c') if is_achievement else None,
                    )
                if not is_middle:
                    grade.full_clean()
                grade.save()
                created_count += 1

            messages.success(request, f'{created_count}개의 내신 성적이 등록되었습니다.')
            return redirect('students:student_detail', pk=student_pk)

        except Exception as e:
            messages.error(request, f'성적 저장 중 오류가 발생했습니다: {str(e)}')
            return redirect('grades:internal_grade_bulk_create', student_pk=student_pk)

    if is_middle:
        subjects_all = Subject.objects.filter(is_active=True, school_level='M').order_by('subject_code')
        subjects_list = [{'id': s.pk, 'name': s.name, 'category': s.category} for s in subjects_all]
        context = {'student': student, 'is_middle': True, 'subjects_list': subjects_list}
        template = 'grades/grade_bulk_form_middle.html'
    elif is_2022:
        TYPE_MAP = {'0': 'common', '1': 'general', '2': 'elective', '3': 'fusion'}
        subjects_2022 = Subject.objects.filter(is_active=True, school_level='H', curriculum_year=2022).order_by('subject_code')
        subjects_list = []
        for s in subjects_2022:
            sc = TYPE_MAP.get(s.subject_code[2], '') if len(s.subject_code) == 6 else ''
            subjects_list.append({
                'id': s.pk,
                'name': f'[{s.subject_code}] {s.name}',
                'category': s.category,
                'classification': sc,
            })
        context = {'student': student, 'is_2022': True, 'subjects_list': subjects_list}
        template = 'grades/grade_bulk_form_2022.html'
    else:
        context = {'student': student, 'is_2022': False}
        template = 'grades/grade_bulk_form.html'
    return render(request, template, context)


@login_required
def get_subjects_by_category(request):
    """교과별 과목 목록 반환 (AJAX)"""
    category = request.GET.get('category', '')
    curriculum_year = request.GET.get('curriculum_year')

    qs = Subject.objects.filter(is_active=True, school_level='H').order_by('subject_code')
    if curriculum_year and curriculum_year.isdigit():
        qs = qs.filter(curriculum_year=int(curriculum_year))

    TYPE_MAP = {'0': 'common', '1': 'general', '2': 'elective', '3': 'fusion'}

    def to_dict(s):
        # 2022 과목(6자리 코드)은 subject_classification을 코드에서 자동 추출
        sc = ''
        if s.curriculum_year == 2022 and len(s.subject_code) == 6:
            sc = TYPE_MAP.get(s.subject_code[2], '')
        return {'id': s.id, 'name': f'[{s.subject_code}] {s.name}', 'classification': sc}

    if category:
        filtered_subjects = [s for s in qs if s.category == category]
        data = [to_dict(s) for s in filtered_subjects]
    else:
        data = [to_dict(s) for s in qs]

    return JsonResponse({'subjects': data})


@login_required
def grade_import(request, student_pk):
    """CSV/Excel 파일에서 성적 일괄 임포트"""
    student = get_object_or_404(Student, pk=student_pk)

    if request.method == 'POST':
        form = GradeImportForm(request.POST, request.FILES)
        if form.is_valid():
            grade_type = form.cleaned_data['grade_type']
            uploaded_file = request.FILES['file']

            # 파일 확장자 확인
            file_name = uploaded_file.name.lower()

            try:
                if file_name.endswith('.csv'):
                    result = process_csv_file(uploaded_file, student, grade_type)
                elif file_name.endswith(('.xlsx', '.xls')):
                    result = process_excel_file(uploaded_file, student, grade_type)
                else:
                    messages.error(request, '지원하지 않는 파일 형식입니다. CSV 또는 Excel 파일만 업로드 가능합니다.')
                    return redirect('grades:grade_import', student_pk=student_pk)

                if result['success']:
                    messages.success(request, f"{result['created_count']}개의 성적이 등록되었습니다.")
                    if result['errors']:
                        for error in result['errors'][:5]:  # 최대 5개 에러만 표시
                            messages.warning(request, error)
                        if len(result['errors']) > 5:
                            messages.warning(request, f"... 외 {len(result['errors']) - 5}개의 오류가 더 있습니다.")
                else:
                    messages.error(request, result['message'])
                    for error in result['errors'][:5]:
                        messages.error(request, error)

                return redirect('students:student_detail', pk=student_pk)

            except Exception as e:
                messages.error(request, f'파일 처리 중 오류가 발생했습니다: {str(e)}')
                return redirect('grades:grade_import', student_pk=student_pk)
    else:
        form = GradeImportForm()

    is_middle = student.grade in ['K7', 'K8', 'K9']
    is_2022 = (not is_middle) and (student.curriculum_year == 2022)
    context = {
        'form': form,
        'student': student,
        'is_middle': is_middle,
        'is_2022': is_2022,
    }
    return render(request, 'grades/grade_import.html', context)


def process_csv_file(uploaded_file, student, grade_type):
    """CSV 파일 처리"""
    errors = []
    created_count = 0

    try:
        # 파일 내용 읽기 (UTF-8 또는 CP949 시도)
        content = uploaded_file.read()
        try:
            decoded_content = content.decode('utf-8-sig')  # BOM 처리
        except UnicodeDecodeError:
            decoded_content = content.decode('cp949')  # 한글 Windows 인코딩

        reader = csv.DictReader(io.StringIO(decoded_content))

        with transaction.atomic():
            for row_num, row in enumerate(reader, start=2):  # 헤더 다음 행부터
                try:
                    grade = create_grade_from_row(row, student, grade_type, row_num)
                    if grade:
                        created_count += 1
                except ValueError as e:
                    errors.append(f"행 {row_num}: {str(e)}")
                except Exception as e:
                    errors.append(f"행 {row_num}: 처리 실패 - {str(e)}")

        return {
            'success': True,
            'created_count': created_count,
            'errors': errors,
            'message': '파일 처리가 완료되었습니다.'
        }

    except Exception as e:
        return {
            'success': False,
            'created_count': 0,
            'errors': errors,
            'message': f'CSV 파일 읽기 실패: {str(e)}'
        }


def process_excel_file(uploaded_file, student, grade_type):
    """Excel 파일 처리"""
    errors = []
    created_count = 0

    try:
        import openpyxl
    except ImportError:
        return {
            'success': False,
            'created_count': 0,
            'errors': [],
            'message': 'Excel 파일 처리를 위해 openpyxl 라이브러리가 필요합니다. pip install openpyxl'
        }

    try:
        workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = workbook.active

        # 헤더 행 읽기
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.strip() if cell.value else '')

        with transaction.atomic():
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # 빈 행 건너뛰기
                    continue

                try:
                    row_dict = dict(zip(headers, row))
                    grade = create_grade_from_row(row_dict, student, grade_type, row_num)
                    if grade:
                        created_count += 1
                except ValueError as e:
                    errors.append(f"행 {row_num}: {str(e)}")
                except Exception as e:
                    errors.append(f"행 {row_num}: 처리 실패 - {str(e)}")

        return {
            'success': True,
            'created_count': created_count,
            'errors': errors,
            'message': '파일 처리가 완료되었습니다.'
        }

    except Exception as e:
        return {
            'success': False,
            'created_count': 0,
            'errors': errors,
            'message': f'Excel 파일 읽기 실패: {str(e)}'
        }


def clean_bom(text):
    """BOM 및 불필요한 문자 제거"""
    if text is None:
        return ''
    text = str(text)
    # 다양한 BOM 문자 제거
    bom_chars = ['\ufeff', '\ufffe', '\xef\xbb\xbf', '\xff\xfe', '\xfe\xff']
    for bom in bom_chars:
        text = text.replace(bom, '')
    return text.strip()


def normalize_row(row):
    """행의 키와 값을 정규화 (BOM 제거, 공백 제거)"""
    normalized = {}
    for key, value in row.items():
        if key:
            # 키에서 BOM과 공백 제거
            clean_key = clean_bom(str(key)).replace(' ', '')
            # 값에서도 BOM 제거
            clean_value = clean_bom(value) if value is not None else value
            normalized[clean_key] = clean_value
    return normalized


def get_value(row, *keys):
    """여러 가능한 키 이름으로 값 찾기"""
    for key in keys:
        if key in row and row[key] is not None:
            return row[key]
    return ''


def create_grade_from_row(row, student, grade_type, row_num):
    """행 데이터에서 Grade 객체 생성"""

    # 행 키 정규화
    row = normalize_row(row)

    # 필수 필드 확인
    subject_code = str(get_value(row, '과목코드', '과목_코드', 'subject_code') or '').strip()
    # '과목' 컬럼(새 형식) 또는 '과목명' 컬럼(구 형식) 모두 지원; '교과'는 카테고리명이므로 과목 찾기에 사용 안 함
    subject_name = str(get_value(row, '과목명', '과목', '과목이름', 'subject_name', 'subject') or '').strip()

    # 일부 성적표 형식은 진로선택 과목의 경우 '과목명'을 비워두고 '진로선택' 컬럼에
    # 과목명을 직접 적어두는 경우가 있다 (예: 과목명='', 진로선택='생명과학2').
    # 이 컬럼 값이 단순 여부 표시(진로선택/예/1 등)가 아니라면 과목명으로 간주한다.
    force_elective = False
    if not subject_name:
        elective_raw = str(get_value(row, '진로선택', '진로_선택', '선택과목', 'elective') or '').strip()
        elective_marker_keywords = {'1', 'true', 'yes', 'y', 'o', '예', '진로선택', '○', 'v', '선택'}
        if elective_raw and elective_raw.lower() not in elective_marker_keywords:
            subject_name = elective_raw
            force_elective = True

    if not subject_code and not subject_name:
        return None  # 빈 행 건너뛰기

    # 과목 찾기
    # 동일한 이름의 과목이 2015/2022 교육과정 간에, 또는 중학교/고등학교 간에
    # 공존할 수 있으므로 (예: '문학'은 2015/2022 고교 과목끼리, '수학'은 고교/중학교끼리
    # 이름이 겹침) 학생의 교육과정·학교급과 일치하는 과목을 우선 선택한다.
    student_curriculum = getattr(student, 'curriculum_year', None)
    student_school_level = 'M' if student.grade in ['K7', 'K8', 'K9'] else 'H'

    def pick_best(candidates):
        candidates = list(candidates)
        if not candidates:
            return None
        for c in candidates:
            if c.school_level == student_school_level and c.curriculum_year == student_curriculum:
                return c
        for c in candidates:
            if c.school_level == student_school_level:
                return c
        if student_curriculum:
            for c in candidates:
                if c.curriculum_year == student_curriculum:
                    return c
        return candidates[0]

    subject = None
    if subject_code:
        subject = Subject.objects.filter(subject_code=subject_code, is_active=True).first()
    if not subject and subject_name:
        # 정확한 이름 매칭 + 띄어쓰기 무시 매칭(기술가정 = 기술 가정)을 한 풀로 모아
        # 학생 교육과정·학교급과 일치하는 과목을 우선 선택한다. (한 쪽 매칭 방식에서만
        # 찾아지는 후보가 다른 교육과정/학교급 것이면 잘못된 과목이 뽑히므로 분리하지 않는다)
        search_name_no_space = subject_name.replace(' ', '')
        name_candidates = [
            s for s in Subject.objects.filter(is_active=True)
            if s.name == subject_name or s.name.replace(' ', '') == search_name_no_space
        ]
        subject = pick_best(name_candidates)
        # 부분 매칭 시도
        if not subject:
            subject = pick_best(Subject.objects.filter(name__icontains=subject_name, is_active=True))

    if not subject:
        raise ValueError(f"과목을 찾을 수 없습니다: 코드={subject_code}, 이름={subject_name}")

    # 공통 필드 파싱 (다양한 헤더명 지원)
    year = parse_int(get_value(row, '학년', 'year', 'grade'), '학년')

    # 과목 분류 파싱 (진로선택/융합선택/공통/일반선택)
    classification_raw = str(get_value(row, '과목분류', '과목_분류', 'subject_classification') or '').strip()
    classification_map = {
        '공통': 'common', '공통과목': 'common',
        '일반': 'general', '일반선택': 'general',
        '진로': 'elective', '진로선택': 'elective',
        '융합': 'fusion', '융합선택': 'fusion',
    }
    subject_classification = classification_map.get(classification_raw, '')

    # '진로선택' 컬럼 값을 과목명으로 사용한 경우, 그 컬럼 자체가 진로선택 표시이므로 확정
    if force_elective:
        subject_classification = 'elective'

    # 2022 교육과정 과목: 과목코드 3번째 자리에서 자동 추출 (0=공통, 1=일반선택, 2=진로선택, 3=융합선택)
    if not subject_classification and subject and len(subject.subject_code) == 6 and subject.curriculum_year == 2022:
        _type_map = {'0': 'common', '1': 'general', '2': 'elective', '3': 'fusion'}
        subject_classification = _type_map.get(subject.subject_code[2], 'general')

    # 하위 호환: 기존 진로선택 여부 컬럼도 지원
    if not subject_classification:
        is_elective_raw = get_value(row, '진로선택', '진로_선택', '선택과목', 'elective')
        is_elective_legacy = str(is_elective_raw).strip().lower() in ['1', 'true', 'yes', 'y', 'o', '예', '진로선택', '○', 'v', '선택']
        subject_classification = 'elective' if is_elective_legacy else 'general'

    is_elective = subject_classification in ('elective', 'fusion')

    # 중복 체크 (grade_type에 따라 다른 조건)
    if grade_type == 'internal':
        semester = parse_int(get_value(row, '학기', 'semester'), '학기')
        existing = Grade.objects.filter(
            student=student,
            grade_type='internal',
            subject=subject,
            year=year,
            semester=semester
        ).exists()
        if existing:
            raise ValueError(f"이미 등록된 성적입니다: {year}학년 {semester}학기 {subject.name}")
    else:  # mock
        exam_year_val = parse_int(get_value(row, '시험연도', '연도', '시험_연도', 'exam_year', 'year'), '시험연도')
        exam_month_val = parse_int(get_value(row, '시험월', '월', '시험_월', 'exam_month', 'month'), '시험월')
        exam_name_val = str(get_value(row, '시험명', '모의고사명', '시험이름', '시험_명', 'exam_name') or '').strip()
        existing = Grade.objects.filter(
            student=student,
            grade_type='mock',
            subject=subject,
            year=year,
            exam_year=exam_year_val,
            exam_month=exam_month_val,
            exam_name=exam_name_val
        ).exists()
        if existing:
            raise ValueError(f"이미 등록된 성적입니다: {exam_year_val}년 {exam_month_val}월 {exam_name_val} {subject.name}")

    score = parse_decimal(get_value(row, '원점수', '점수', 'score'), '원점수')
    subject_average_raw = get_value(row, '과목평균', '평균', '과목_평균', 'average', 'avg')
    subject_average = parse_decimal(subject_average_raw, '과목평균') if subject_average_raw and str(subject_average_raw).strip() else None

    # 등급 — 중학교 파일에는 없으므로 선택사항 / 2022 파일은 '상대등급' 컬럼
    grade_rank_raw = get_value(row, '등급', '상대등급', 'rank', 'grade_rank')
    is_middle_format = not (grade_rank_raw and str(grade_rank_raw).strip())

    # 진로선택 과목일 경우 등급/표준편차 대신 성취도/분포비율
    if is_elective and grade_type == 'internal':
        subject_stddev = None
        grade_rank = None
    elif is_middle_format:
        # 중학교 형식: 등급 없음
        subject_stddev = parse_decimal_optional(get_value(row, '표준편차', '표준_편차', 'stddev', 'std'))
        grade_rank = None
    else:
        # 표준편차는 선택사항 (2022 과정 파일에는 없을 수 있음)
        subject_stddev = parse_decimal_optional(get_value(row, '표준편차', '표준_편차', 'stddev', 'std'))
        grade_rank = parse_int(grade_rank_raw, '상대등급')
        # 2022 교육과정 고1·고2: 1~5등급 / 2015 교육과정: 1~9등급
        if subject and subject.curriculum_year == 2022:
            if not (1 <= grade_rank <= 5):
                raise ValueError(f"2022 과정 상대등급은 1~5 사이여야 합니다: {grade_rank}")
        else:
            if not (1 <= grade_rank <= 9):
                raise ValueError(f"등급은 1~9 사이여야 합니다: {grade_rank}")

    # Grade 객체 생성
    grade_obj = Grade(
        student=student,
        grade_type=grade_type,
        subject=subject,
        year=year,
        score=score,
        subject_average=subject_average,
        subject_stddev=subject_stddev,
        grade_rank=grade_rank,
    )

    if grade_type == 'internal':
        # 내신 전용 필드
        semester = parse_int(get_value(row, '학기', 'semester'), '학기')
        # 단위수/학점 (중학교 형식은 없으므로 선택사항)
        credits_raw = get_value(row, '단위수', '단위', '학점', '이수단위', 'credits', 'unit')
        credits = parse_int(credits_raw, '단위수') if (credits_raw and str(credits_raw).strip()) else None

        if semester not in [1, 2]:
            raise ValueError(f"학기는 1 또는 2여야 합니다: {semester}")

        # 수강자수 (중학교 형식)
        enrolled_raw = get_value(row, '수강자수', '수강자_수', 'enrolled', 'enrolled_count')
        enrolled_count = int(float(str(enrolled_raw).strip())) if (enrolled_raw and str(enrolled_raw).strip()) else None

        # 석차, 동석차수, 백분위 (선택사항)
        subject_rank_raw = get_value(row, '석차', 'subject_rank')
        subject_rank = int(float(str(subject_rank_raw).strip())) if (subject_rank_raw and str(subject_rank_raw).strip()) else None
        same_rank_raw = get_value(row, '동석차수', '동석차_수', 'same_rank_count')
        same_rank_count = int(float(str(same_rank_raw).strip())) if (same_rank_raw and str(same_rank_raw).strip()) else None
        percentile_raw = get_value(row, '백분위', 'percentile')
        percentile_val = parse_decimal_optional(percentile_raw)

        grade_obj.semester = semester
        grade_obj.credits = credits
        grade_obj.subject_classification = subject_classification
        grade_obj.is_elective = is_elective
        if enrolled_count is not None:
            grade_obj.enrolled_count = enrolled_count
        if subject_rank is not None:
            grade_obj.subject_rank = subject_rank
        if same_rank_count is not None:
            grade_obj.same_rank_count = same_rank_count
        if percentile_val is not None:
            grade_obj.percentile = percentile_val

        # 성취도 파싱: 진로/융합선택은 A/B/C, 중학교 형식은 A~E
        if is_elective:
            achievement_level = str(get_value(row, '성취도', 'achievement', 'achievement_level') or '').strip().upper()
            if achievement_level not in ['A', 'B', 'C']:
                raise ValueError(f"성취도는 A, B, C 중 하나여야 합니다: {achievement_level}")

            distribution_a = parse_decimal_optional(get_value(row, '분포비율A', 'A비율', '성취도A비율', 'distribution_a'))
            distribution_b = parse_decimal_optional(get_value(row, '분포비율B', 'B비율', '성취도B비율', 'distribution_b'))
            distribution_c = parse_decimal_optional(get_value(row, '분포비율C', 'C비율', '성취도C비율', 'distribution_c'))

            # 2022 교육과정 진로선택은 분포비율 선택사항 (2015는 필수)
            _is_2022_subject = subject and subject.curriculum_year == 2022
            if not _is_2022_subject:
                if distribution_a is None:
                    raise ValueError("진로선택 과목은 분포비율A가 필수입니다")
                if distribution_b is None:
                    raise ValueError("진로선택 과목은 분포비율B가 필수입니다")
                if distribution_c is None:
                    raise ValueError("진로선택 과목은 분포비율C가 필수입니다")

            grade_obj.achievement_level = achievement_level
            grade_obj.distribution_a = distribution_a
            grade_obj.distribution_b = distribution_b
            grade_obj.distribution_c = distribution_c
        elif is_middle_format:
            # 중학교 형식: 성취도 A~E (선택)
            achievement_raw = str(get_value(row, '성취도', 'achievement', 'achievement_level') or '').strip().upper()
            if achievement_raw and achievement_raw in ['A', 'B', 'C', 'D', 'E']:
                grade_obj.achievement_level = achievement_raw

    else:  # mock
        # 모의고사 전용 필드
        exam_year = parse_int(get_value(row, '시험연도', '연도', '시험_연도', 'exam_year', 'year'), '시험연도')
        exam_month = parse_int(get_value(row, '시험월', '월', '시험_월', 'exam_month', 'month'), '시험월')
        exam_name = str(get_value(row, '시험명', '모의고사명', '시험이름', '시험_명', 'exam_name') or '').strip()
        percentile = parse_decimal(get_value(row, '백분위', '백분_위', 'percentile'), '백분위')

        if not exam_name:
            raise ValueError("시험명이 필요합니다")
        if not (1 <= exam_month <= 12):
            raise ValueError(f"시험월은 1~12 사이여야 합니다: {exam_month}")

        grade_obj.exam_year = exam_year
        grade_obj.exam_month = exam_month
        grade_obj.exam_name = exam_name
        grade_obj.percentile = percentile

    if not is_middle_format:
        grade_obj.full_clean()
    grade_obj.save()
    return grade_obj


def parse_int(value, field_name):
    """정수 파싱 헬퍼"""
    if value is None or str(value).strip() == '':
        raise ValueError(f"{field_name}이(가) 필요합니다")
    try:
        return int(float(str(value).strip()))  # Excel에서 숫자가 float로 올 수 있음
    except (ValueError, TypeError):
        raise ValueError(f"{field_name} 값이 올바르지 않습니다: {value}")


def parse_decimal(value, field_name):
    """소수 파싱 헬퍼"""
    if value is None or str(value).strip() == '':
        raise ValueError(f"{field_name}이(가) 필요합니다")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError(f"{field_name} 값이 올바르지 않습니다: {value}")


def parse_decimal_optional(value):
    """선택적 소수 파싱 헬퍼"""
    if value is None or str(value).strip() == '':
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


@login_required
def delete_all_grades(request, student_pk):
    """학생의 모든 성적 삭제"""
    student = get_object_or_404(Student, pk=student_pk)

    if request.method == 'POST':
        deleted_count = Grade.objects.filter(student=student).count()
        Grade.objects.filter(student=student).delete()
        messages.success(request, f'{deleted_count}개의 성적이 삭제되었습니다.')

    return redirect('students:student_detail', pk=student_pk)


@login_required
def student_grades(request, student_pk):
    """학생 성적 전용 페이지"""
    student = get_object_or_404(Student, pk=student_pk)

    # 성적 데이터 조회
    internal_grades = Grade.objects.filter(
        student=student,
        grade_type='internal'
    ).select_related('subject').order_by('-year', '-semester', 'subject__subject_code')

    mock_grades = Grade.objects.filter(
        student=student,
        grade_type='mock'
    ).select_related('subject').order_by('-exam_year', '-exam_month', 'subject__subject_code')

    # 학기별 평균 내신 등급 계산 (진로선택 과목 제외)
    semester_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})
    year_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})

    # 학기별/교과별 성적 데이터 (차트용) - 진로선택 제외, 교과별 집계
    semester_category_grades = defaultdict(lambda: defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0}))

    # 교과별 전체 통계 (교과 조합 분석용)
    category_stats = defaultdict(lambda: {'total_weighted': 0, 'total_credits': 0})

    for grade in internal_grades:
        # 등급/단위수가 없는 성적(중학교 등)은 통계 계산에서 제외
        if grade.grade_rank is None or grade.credits is None:
            continue

        is_achievement = grade.subject_classification in ('elective', 'fusion')
        # 차트용 데이터 수집 (진로선택/융합선택 제외)
        if not is_achievement:
            key = f"{grade.year}-{grade.semester}"
            category = grade.curriculum or '기타'
            semester_category_grades[key][category]['total_weighted'] += grade.grade_rank * grade.credits
            semester_category_grades[key][category]['total_credits'] += grade.credits

            # 교과별 전체 통계
            category_stats[category]['total_weighted'] += grade.grade_rank * grade.credits
            category_stats[category]['total_credits'] += grade.credits

        if is_achievement:  # 진로선택/융합선택 과목은 평균 계산에서 제외
            continue

        semester_key = (grade.year, grade.semester)
        semester_stats[semester_key]['total_weighted'] += grade.grade_rank * grade.credits
        semester_stats[semester_key]['total_credits'] += grade.credits

        # 학년별 통계
        year_stats[grade.year]['total_weighted'] += grade.grade_rank * grade.credits
        year_stats[grade.year]['total_credits'] += grade.credits

    # 학기별 평균 계산 및 정렬
    semester_averages = []
    for (year, semester), stats in sorted(semester_stats.items()):
        if stats['total_credits'] > 0:
            avg = Decimal(stats['total_weighted']) / Decimal(stats['total_credits'])
            semester_averages.append({
                'year': year,
                'semester': semester,
                'average': round(avg, 2),
                'total_credits': stats['total_credits'],
            })

    # 전체 평균 등급 계산 (동일 가중치)
    total_weighted_sum = sum(s['total_weighted'] for s in semester_stats.values())
    total_credits_sum = sum(s['total_credits'] for s in semester_stats.values())
    overall_average = None
    if total_credits_sum > 0:
        overall_average = round(Decimal(total_weighted_sum) / Decimal(total_credits_sum), 2)

    # 학년별 가중치 적용 전체 등급 계산
    weighted_averages = []
    weight_configs = [
        {'name': '30:30:40', 'weights': {1: 30, 2: 30, 3: 40}},
        {'name': '20:40:40', 'weights': {1: 20, 2: 40, 3: 40}},
        {'name': '20:30:50', 'weights': {1: 20, 2: 30, 3: 50}},
    ]

    for config in weight_configs:
        weights = config['weights']
        weighted_sum = Decimal(0)
        weight_sum = Decimal(0)

        for year in [1, 2, 3]:
            if year in year_stats and year_stats[year]['total_credits'] > 0:
                year_avg = Decimal(year_stats[year]['total_weighted']) / Decimal(year_stats[year]['total_credits'])
                weighted_sum += year_avg * Decimal(weights[year])
                weight_sum += Decimal(weights[year])

        if weight_sum > 0:
            weighted_avg = round(weighted_sum / weight_sum, 2)
            weighted_averages.append({
                'name': config['name'],
                'average': weighted_avg,
            })

    # 차트용 데이터 준비 (교과별 평균 등급)
    chart_data = []
    for semester_key in sorted(semester_category_grades.keys()):
        year, sem = semester_key.split('-')
        semester_data = {
            'label': f"{year}학년 {sem}학기",
            'categories': {}
        }
        for category, stats in semester_category_grades[semester_key].items():
            if stats['total_credits'] > 0:
                avg_grade = round(float(stats['total_weighted']) / float(stats['total_credits']), 2)
                semester_data['categories'][category] = {
                    'average': avg_grade,
                    'total_credits': stats['total_credits'],
                }
        chart_data.append(semester_data)

    # 과목별 등급 추이 (꺾은선 그래프용)
    subject_grades_map = {}  # {subject_name: {'category': str, 'data': {sem_label: grade_rank}}}
    all_sems_ordered = []
    seen_sems = set()
    for grade in sorted(internal_grades, key=lambda g: (g.year, g.semester)):
        if grade.grade_rank is None:
            continue
        if grade.subject_classification in ('elective', 'fusion'):
            continue
        sem_label = f"{grade.year}학년 {grade.semester}학기"
        if sem_label not in seen_sems:
            all_sems_ordered.append(sem_label)
            seen_sems.add(sem_label)
        sub_name = grade.subject.name if grade.subject else '기타'
        if sub_name not in subject_grades_map:
            subject_grades_map[sub_name] = {'category': grade.curriculum or '기타', 'data': {}}
        subject_grades_map[sub_name]['data'][sem_label] = float(grade.grade_rank)
    subject_trend_json = json.dumps({
        'semesters': all_sems_ordered,
        'subjects': [
            {
                'name': name,
                'category': info['category'],
                'data': [info['data'].get(sem) for sem in all_sems_ordered],
            }
            for name, info in subject_grades_map.items()
        ],
    }, ensure_ascii=False)

    # 교과 조합별 평균 분석
    category_combinations = [
        {'name': '국수영과', 'categories': ['국어', '수학', '영어', '과학']},
        {'name': '국수영사', 'categories': ['국어', '수학', '영어', '사회']},
        {'name': '국수영사과', 'categories': ['국어', '수학', '영어', '사회', '과학']},
    ]

    combination_averages = []
    for combo in category_combinations:
        total_weighted = 0
        total_credits = 0
        missing_categories = []

        for cat in combo['categories']:
            if cat in category_stats and category_stats[cat]['total_credits'] > 0:
                total_weighted += category_stats[cat]['total_weighted']
                total_credits += category_stats[cat]['total_credits']
            else:
                missing_categories.append(cat)

        if total_credits > 0:
            avg = round(Decimal(total_weighted) / Decimal(total_credits), 2)
            combination_averages.append({
                'name': combo['name'],
                'categories': combo['categories'],
                'average': avg,
                'total_credits': total_credits,
                'missing': missing_categories,
            })

    # 일반 내신 성적과 진로선택/융합선택 성적 분리
    regular_internal_grades = [g for g in internal_grades if g.subject_classification not in ('elective', 'fusion')]
    elective_grades = [g for g in internal_grades if g.subject_classification in ('elective', 'fusion')]

    # ── 진도 평가 분석 ──────────────────────────────────────────
    from bookstore.models import BookSale
    from progress.models import LearningRecord

    ACHIEVEMENT_META = [
        {'code': 'A', 'label': 'A (우수)', 'color': 'indigo'},
        {'code': 'B', 'label': 'B (양호)', 'color': 'green'},
        {'code': 'C', 'label': 'C (보통)', 'color': 'yellow'},
        {'code': 'D', 'label': 'D (미흡)', 'color': 'orange'},
        {'code': 'F', 'label': 'F (재학습)', 'color': 'red'},
    ]

    all_sales = BookSale.objects.filter(student=student).select_related('book').order_by('-sale_date', '-pk')

    book_progress_data = []
    overall_counts = {m['code']: 0 for m in ACHIEVEMENT_META}

    for sale in all_sales:
        records = LearningRecord.objects.filter(
            student=student,
            book_sale=sale,
            record_type='textbook',
            achievement__in=['A', 'B', 'C', 'D', 'F']
        ).select_related('book_content').order_by('book_content__chapter_num', 'book_content__subsection_num')

        if not records.exists():
            continue

        counts = {m['code']: 0 for m in ACHIEVEMENT_META}
        pages_by_level = {m['code']: [] for m in ACHIEVEMENT_META}

        for rec in records:
            counts[rec.achievement] += 1
            overall_counts[rec.achievement] += 1
            if rec.book_content:
                pages_by_level[rec.achievement].append({
                    'page': rec.book_content.page,
                    'title': rec.book_content.subsection_title,
                    'section': rec.book_content.section_title or '',
                    'study_date': rec.date,
                })

        total = sum(counts.values())
        levels = []
        for m in ACHIEVEMENT_META:
            cnt = counts[m['code']]
            levels.append({
                'code': m['code'],
                'label': m['label'],
                'color': m['color'],
                'count': cnt,
                'percent': round(cnt / total * 100, 1) if total else 0,
                'pages': pages_by_level[m['code']],
            })

        segments_json = json.dumps([
            {'p': l['percent'], 'c': l['color']}
            for l in levels if l['percent'] > 0
        ])
        book_progress_data.append({
            'sale': sale,
            'book': sale.book,
            'total': total,
            'levels': levels,
            'is_completed': sale.is_learning_completed,
            'segments_json': segments_json,
        })

    overall_total = sum(overall_counts.values())
    overall_levels = []
    for m in ACHIEVEMENT_META:
        cnt = overall_counts[m['code']]
        overall_levels.append({
            'code': m['code'],
            'label': m['label'],
            'color': m['color'],
            'count': cnt,
            'percent': round(cnt / overall_total * 100, 1) if overall_total else 0,
        })
    overall_segments_json = json.dumps([
        {'p': l['percent'], 'c': l['color']}
        for l in overall_levels if l['percent'] > 0
    ])

    # ── 퀴즈/테스트 기록 분석 ────────────────────────────────────
    TEST_TYPE_CHOICES = [
        ('quiz', '퀴즈'),
        ('practice', '연습 문제'),
        ('booklet', '제본 교재'),
        ('entrance_exam', '입학 시험'),
        ('other', '기타'),
    ]
    test_records = LearningRecord.objects.filter(
        student=student,
        record_type__in=['quiz', 'practice', 'booklet', 'entrance_exam', 'other']
    ).select_related('subject', 'teacher').order_by('-date', '-created_at')

    test_counts = {m['code']: 0 for m in ACHIEVEMENT_META}
    for rec in test_records:
        if rec.achievement in test_counts:
            test_counts[rec.achievement] += 1
    test_total_evaluated = sum(test_counts.values())
    test_levels = []
    for m in ACHIEVEMENT_META:
        cnt = test_counts[m['code']]
        test_levels.append({
            'code': m['code'],
            'label': m['label'],
            'color': m['color'],
            'count': cnt,
            'percent': round(cnt / test_total_evaluated * 100, 1) if test_total_evaluated else 0,
        })
    test_segments_json = json.dumps([
        {'p': l['percent'], 'c': l['color']}
        for l in test_levels if l['percent'] > 0
    ])

    # 각 기록에 quiz_detail 첨부
    processed_test_records = []
    for rec in test_records:
        quiz_detail = None
        if rec.quiz_results:
            total = rec.quiz_results.get('total', 0)
            wrong_nums = sorted(rec.quiz_results.get('wrong', []))
            correct_count = total - len(wrong_nums)
            correct_pct = round(correct_count / total * 100, 1) if total else 0
            wrong_pct = round(100 - correct_pct, 1) if total else 0
            quiz_detail = {
                'total': total,
                'wrong_nums': wrong_nums,  # sorted list for display
                'q_range': list(range(1, total + 1)),
                'correct_count': correct_count,
                'wrong_count': len(wrong_nums),
                'segments_json': json.dumps([
                    {'p': correct_pct, 'c': '#22c55e'},
                    {'p': wrong_pct, 'c': '#f87171'},
                ]) if total > 0 else '[]',
            }
        processed_test_records.append({'rec': rec, 'quiz_detail': quiz_detail})
    # ────────────────────────────────────────────────────────────

    context = {
        'student': student,
        'internal_grades': regular_internal_grades,
        'elective_grades': elective_grades,
        'mock_grades': mock_grades,
        'semester_averages': semester_averages,
        'overall_average': overall_average,
        'weighted_averages': weighted_averages,
        'combination_averages': combination_averages,
        'chart_data': json.dumps(chart_data, ensure_ascii=False),
        'subject_trend_json': subject_trend_json,
        # 진도 평가 분석
        'book_progress_data': book_progress_data,
        'overall_levels': overall_levels,
        'overall_total': overall_total,
        'overall_segments_json': overall_segments_json if overall_total > 0 else '[]',
        'achievement_meta': ACHIEVEMENT_META,
        # 퀴즈/테스트 기록
        'test_records': processed_test_records,
        'test_levels': test_levels,
        'test_total_evaluated': test_total_evaluated,
        'test_segments_json': test_segments_json,
        'test_type_choices': TEST_TYPE_CHOICES,
        'subjects': Subject.objects.filter(is_active=True).annotate(
            _math_first=Case(
                When(subject_code__startswith='2', then=Value(0)),
                default=Value(1),
                output_field=IntegerField()
            )
        ).order_by('_math_first', 'subject_code'),
        'is_middle': student.grade in ['K7', 'K8', 'K9'],
        'middle_grades_json': json.dumps([
            {
                'id': g.pk,
                'subject': g.subject.name,
                'year': g.year,
                'semester': g.semester,
                'score': float(g.score),
                'average': float(g.subject_average),
                'stddev': float(g.subject_stddev),
                'enrolled': g.enrolled_count,
                'achievement': g.achievement_level or '',
            }
            for g in internal_grades
            if g.score is not None
            and g.subject_average is not None
            and g.subject_stddev is not None
            and float(g.subject_stddev) > 0
        ], ensure_ascii=False) if student.grade in ['K7', 'K8', 'K9'] else '[]',
    }
    return render(request, 'grades/student_grades.html', context)


@login_required
def test_record_create(request, student_pk):
    """퀴즈/테스트 기록 생성"""
    from progress.models import LearningRecord
    from teachers.models import Teacher

    student = get_object_or_404(Student, pk=student_pk)
    if request.method == 'POST':
        record_type = request.POST.get('record_type', 'quiz')
        title = request.POST.get('title', '').strip()
        date = request.POST.get('date') or None
        achievement = request.POST.get('achievement', '')
        score_str = request.POST.get('score', '').strip()
        total_score_str = request.POST.get('total_score', '').strip()
        needs_review = request.POST.get('needs_review') == 'on'
        memo = request.POST.get('memo', '').strip()
        subject_pk = request.POST.get('subject', '')

        if not title:
            messages.error(request, '제목을 입력해 주세요.')
            return redirect('grades:student_grades', student_pk=student_pk)

        # 문제별 정답/오답 상세 처리
        quiz_total_str = request.POST.get('quiz_total', '').strip()
        quiz_wrong_str = request.POST.get('quiz_wrong', '').strip()
        quiz_results_data = None
        if quiz_total_str and quiz_total_str.isdigit():
            q_total = int(quiz_total_str)
            wrong_list = sorted([int(x) for x in quiz_wrong_str.split(',') if x.strip().isdigit()]) if quiz_wrong_str else []
            quiz_results_data = {'total': q_total, 'wrong': wrong_list}
            if not score_str:
                score_str = str(q_total - len(wrong_list))
                total_score_str = str(q_total)

        score = Decimal(score_str) if score_str else None
        total_score = Decimal(total_score_str) if total_score_str else None
        subject = Subject.objects.filter(pk=subject_pk).first() if subject_pk else None

        teacher = None
        if hasattr(request.user, 'teacher_profile'):
            teacher = request.user.teacher_profile

        LearningRecord.objects.create(
            student=student,
            record_type=record_type,
            title=title,
            date=date,
            achievement=achievement,
            score=score,
            total_score=total_score,
            needs_review=needs_review,
            memo=memo,
            subject=subject,
            teacher=teacher,
            quiz_results=quiz_results_data,
        )
        messages.success(request, f'기록이 추가되었습니다.')
    return redirect('grades:student_grades', student_pk=student_pk)


@login_required
def test_record_delete(request, pk):
    """퀴즈/테스트 기록 삭제"""
    from progress.models import LearningRecord
    record = get_object_or_404(LearningRecord, pk=pk)
    student_pk = record.student.pk
    if request.method == 'POST':
        record.delete()
        messages.success(request, '기록이 삭제되었습니다.')
    return redirect('grades:student_grades', student_pk=student_pk)


@login_required
def test_record_update(request, pk):
    """퀴즈/테스트 기록 수정"""
    from progress.models import LearningRecord
    record = get_object_or_404(LearningRecord, pk=pk)
    student_pk = record.student.pk
    if request.method == 'POST':
        record_type = request.POST.get('record_type', record.record_type)
        title = request.POST.get('title', '').strip()
        date = request.POST.get('date') or None
        achievement = request.POST.get('achievement', '')
        score_str = request.POST.get('score', '').strip()
        total_score_str = request.POST.get('total_score', '').strip()
        memo = request.POST.get('memo', '').strip()
        subject_pk = request.POST.get('subject', '')

        if not title:
            messages.error(request, '제목을 입력해 주세요.')
            return redirect('grades:student_grades', student_pk=student_pk)

        quiz_total_str = request.POST.get('quiz_total', '').strip()
        quiz_wrong_str = request.POST.get('quiz_wrong', '').strip()
        quiz_results_data = None
        if quiz_total_str and quiz_total_str.isdigit():
            q_total = int(quiz_total_str)
            wrong_list = sorted([int(x) for x in quiz_wrong_str.split(',') if x.strip().isdigit()]) if quiz_wrong_str else []
            quiz_results_data = {'total': q_total, 'wrong': wrong_list}
            if not score_str:
                score_str = str(q_total - len(wrong_list))
                total_score_str = str(q_total)

        record.record_type = record_type
        record.title = title
        record.date = date
        record.achievement = achievement
        record.score = Decimal(score_str) if score_str else None
        record.total_score = Decimal(total_score_str) if total_score_str else None
        record.memo = memo
        record.subject = Subject.objects.filter(pk=subject_pk).first() if subject_pk else None
        record.quiz_results = quiz_results_data
        record.save()
        messages.success(request, '기록이 수정되었습니다.')
    return redirect('grades:student_grades', student_pk=student_pk)


def download_grade_template(request, template_type):
    # 관리자 로그인 또는 학부모 세션 중 하나면 허용
    if not request.user.is_authenticated and not request.session.get('parent_student_id'):
        from django.shortcuts import redirect
        return redirect('login')

    fmt = request.GET.get('format', 'xlsx')  # 기본값 xlsx

    if template_type == 'middle':
        base_filename = 'middle_grade_template'
        headers = ['학년', '학기', '과목명', '원점수', '과목평균', '표준편차', '수강자수', '성취도']
        sample_data = [
            [1, 1, '국어', 85.0, 72.0, 15.0, 150, 'B'],
            [1, 1, '수학', 92.0, 68.5, 18.0, 150, 'A'],
            [1, 1, '영어', 78.0, 71.0, 13.5, 150, 'B'],
            [1, 1, '과학', 88.0, 74.2, 12.0, 148, 'A'],
            [1, 1, '사회', 75.0, 70.5, 14.0, 150, 'C'],
        ]
    elif template_type == 'internal_2022':
        base_filename = 'internal_grade_template_2022'
        headers = ['학년', '학기', '교과', '과목', '단위수', '점수', '성취도', '석차', '동석차수', '수강자수', '백분위', '상대등급']
        sample_data = [
            # 공통과목/일반선택: 상대등급(1~5) 입력, 성취도 비워둠
            [1, 1, '국어', '공통국어1', 4, 85, '', 5, 2, 200, 97.5, 2],
            [1, 1, '수학', '공통수학1', 4, 92, '', 3, 1, 200, 98.5, 1],
            [1, 1, '영어', '영어1', 4, 88, '', 8, 3, 200, 96.0, 2],
            [1, 1, '수학', '대수', 4, 90, '', 6, 2, 180, 96.7, 1],
            # 진로선택: 성취도(A/B/C) 입력, 상대등급 비워둠
            [1, 1, '과학', '역학과 에너지', 2, 78, 'A', 12, 5, 120, 90.0, ''],
            [1, 1, '사회', '한국지리 탐구', 2, 82, 'B', 15, 4, 130, 88.5, ''],
        ]
    elif template_type == 'internal':
        base_filename = 'internal_grade_template'
        headers = ['학년', '학기', '과목명', '단위', '원점수', '과목평균', '표준편차', '등급', '진로선택', '성취도', '분포비율A', '분포비율B', '분포비율C']
        sample_data = [
            [1, 1, '국어', 3, 85, 70.5, 12.3, 2, '', '', '', '', ''],
            [1, 1, '수학Ⅰ', 4, 92, 68.2, 15.1, 1, '', '', '', '', ''],
            [1, 1, '영어', 3, 88, 72.1, 11.5, 2, '', '', '', '', ''],
            [2, 1, '물리학Ⅱ', 3, 85, 72.3, '', '', '진로선택', 'A', 25.5, 45.2, 29.3],
            [2, 1, '화학Ⅱ', 3, 78, 68.5, '', '', '진로선택', 'B', 20.1, 50.3, 29.6],
        ]
    else:  # mock
        base_filename = 'mock_exam_template'
        headers = ['학년', '시험연도', '시험월', '시험명', '과목명', '원점수', '과목평균', '표준편차', '등급', '백분위']
        sample_data = [
            [2, 2024, 3, '3월 학력평가', '국어', 85, 70.5, 12.3, 2, 88],
            [2, 2024, 3, '3월 학력평가', '수학', 92, 68.2, 15.1, 1, 95],
            [2, 2024, 6, '6월 모의평가', '국어', 82, 68.1, 11.8, 2, 85],
        ]

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '성적 입력'

        # 헤더 행 스타일
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(h) * 2.5, 12)

        # 데이터 행
        for row in sample_data:
            ws.append(row)

        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{base_filename}.xlsx"'
        return response

    # CSV 응답
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{base_filename}.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in sample_data:
        writer.writerow(row)

    return response
